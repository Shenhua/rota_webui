
"""
Rota Optimizer — v2.9
=====================
Additions vs v2.8:
- **Matrice**: adds an aligned summary table (below) showing **# personnes** par type de poste (Jour/Soir/Nuit/Admin) pour chaque jour.
- **Optional team borders**: `--matrice-team-borders` draws a thin black rectangle around blocks of people **grouped by workdays_per_week** (e.g., 4j vs 3j).
- **Per-person sheets**: adds a large **name header**, then a smaller **identity panel** (équipe, % temps, préférences, EDO, etc.), then the weekly table.

Other behavior (multi-seed `--tries`, scoring, spinner, -v/-vv) unchanged from v2.8.
"""

import argparse, random, itertools as it, sys, time, threading, os
import pandas as pd
from statistics import pstdev
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.formatting.rule import CellIsRule

# ------------------------------
# Constants & visuals
# ------------------------------

JOURS  = ["Lun","Mar","Mer","Jeu","Ven"]
HEURES = {"D":10,"E":10,"N":12,"A":8}
CODE   = {"D":"J","E":"S","N":"N","A":"A"}
REVCODE= {"J":"D","S":"E","N":"N","A":"A"}

COLORS = {"J":"DDEEFF","S":"FFE4CC","N":"E6CCFF","A":"DDDDDD","OFF":"EEEEEE","EDO":"D8D8D8","EDO*":"FFC7CE"}

THIN    = Side(border_style="thin",   color="CCCCCC")
MEDIUM  = Side(border_style="medium", color="BBBBBB")
DOUBLEK = Side(border_style="double", color="000000")
BLACKTHIN = Side(border_style="thin", color="000000")
BORDER_THIN = Border(top=THIN,bottom=THIN,left=THIN,right=THIN)

def week_range(W): 
    return list(range(1, W+1))

# ------------------------------
# Minimal terminal UI (spinner + log + box)
# ------------------------------

SPIN_FRAMES = "⠋⠙⠹⠸⠼⠴⠦⠧⠇⠏"
def _supports_color():
    if os.environ.get("NO_COLOR"): return False
    return sys.stdout.isatty()

def _c(code): 
    return f"\033[{code}m" if _supports_color() else ""

C_RESET = _c("0")
C_DIM   = _c("2")
C_BOLD  = _c("1")
C_OK    = _c("32")
C_WARN  = _c("33")
C_ERR   = _c("31")
C_INFO  = _c("36")
C_TITLE = _c("95")

class Spinner:
    def __init__(self, text, enabled=True):
        self.text = text
        self.enabled = enabled and sys.stdout.isatty()
        self._stop = threading.Event()
        self._thr = None
        self._start_t = None

    def __enter__(self):
        if not self.enabled: 
            print(f"{C_DIM}.. {self.text}{C_RESET}")
            self._start_t = time.time()
            return self
        self._start_t = time.time()
        self._thr = threading.Thread(target=self._run, daemon=True)
        self._thr.start()
        return self

    def _run(self):
        i=0
        while not self._stop.is_set():
            frame = SPIN_FRAMES[i % len(SPIN_FRAMES)]
            sys.stdout.write(f"\r{C_DIM}{frame}{C_RESET} {self.text}")
            sys.stdout.flush()
            time.sleep(0.08)
            i += 1

    def __exit__(self, exc_type, exc, tb):
        elapsed = time.time() - self._start_t
        if self.enabled:
            sys.stdout.write("\r")
        if exc is None:
            print(f"{C_OK}✔{C_RESET} {self.text} {C_DIM}({elapsed:.2f}s){C_RESET}")
        else:
            print(f"{C_ERR}✖{C_RESET} {self.text} {C_DIM}({elapsed:.2f}s){C_RESET}")
        self._stop.set()
        if self._thr:
            self._thr.join(timeout=0.1)

def box(title, lines):
    if not sys.stdout.isatty():
        print(f"\n{title}\n" + "\n".join(lines))
        return
    w = max(len(title), *(len(s) for s in lines)) + 2
    top    = f"┏{'━'*w}┓"
    tline  = f"┃ {C_TITLE}{title:{w-2}s}{C_RESET} ┃"
    sep    = f"┣{'━'*w}┫"
    bottom = f"┗{'━'*w}┛"
    print("\n"+top)
    print(tline)
    print(sep)
    for s in lines:
        print(f"┃ {s:{w-2}s} ┃")
    print(bottom+"\n")

def vlog(verbose, needed, msg):
    if verbose >= needed:
        print(f"{C_DIM}   ↳ {msg}{C_RESET}")

# ------------------------------
# Input & helpers
# ------------------------------

def load_people(path):
    df = pd.read_csv(path).fillna("")
    for c in ["name","workdays_per_week","weeks_pattern"]:
        if c not in df:
            raise SystemExit(f"CSV missing column: {c}")
    for c in ["prefers_night","no_evening","max_nights","edo_eligible","edo_fixed_day","team"]:
        if c not in df: 
            df[c] = ""

    def toint(x, d=0):
        try:
            return int(x)
        except:
            return d

    P = []
    for _,r in df.iterrows():
        name = str(r["name"]).strip()
        if not name:
            continue
        edof = str(r["edo_fixed_day"]).strip()
        team = str(r["team"]).strip() if "team" in r and str(r["team"]).strip() else ""
        P.append({
            "n":   name,
            "wd":  toint(r["workdays_per_week"],0),
            "pat": max(1,toint(r["weeks_pattern"],1)),
            "pn":  toint(r["prefers_night"],0),
            "ne":  toint(r["no_evening"],0),
            "mxn": toint(r["max_nights"],10**6),
            "edo": toint(r["edo_eligible"],0),
            "edof": edof if edof in JOURS else "",
            "team": team,  # optional free text; if empty we fallback to wd-based team labels
        })
    return P

def make_cohorts_fairness(P, mode):
    if mode != "by-wd":
        names = [p["n"] for p in P]
        return {"all": names}, {n:"all" for n in names}
    cohorts, c_of = {}, {}
    for p in P:
        cid = f"{p['wd']}j"
        cohorts.setdefault(cid, []).append(p["n"])
        c_of[p["n"]] = cid
    return cohorts, c_of

def build_edo_plan(P, W, fixed_global=None):
    plan = {w:set() for w in week_range(W)}
    groups = {}
    for p in P:
        if p["edo"]:
            groups.setdefault(p["wd"], []).append(p["n"])
    for arr in groups.values():
        arr = sorted(arr); half = (len(arr)+1)//2
        for w in week_range(W):
            take = arr[:half] if (w % 2) else arr[half:]
            plan[w].update(take)
    fixed = {p["n"]: (p["edof"] or fixed_global or "") for p in P}
    return plan, fixed

def derive_staffing(P, W, edo_plan, verbose=0):
    weeks = {}
    total_wd = sum(p["wd"] for p in P)
    for w in week_range(W):
        edo_count   = sum(1 for p in P if p["edo"] and p["n"] in edo_plan[w])
        person_days = total_wd - edo_count
        admin_days  = 1 if (person_days % 2) else 0
        rem         = person_days - admin_days
        baseN=1; pairsN=baseN*5
        pairs=max(0, (rem - 2*pairsN)//2)
        per_day = {d:{"D":0,"E":0,"N":baseN,"A":0} for d in JOURS}
        if admin_days:
            per_day[JOURS[0]]["A"] = 1
        fill = list(it.product(JOURS, ["D","E"]))
        i=0
        while pairs>0:
            d,s = fill[i % len(fill)]
            per_day[d][s] += 1
            pairs -= 1; i += 1
        weeks[w] = per_day
        vlog(verbose, 1, f"Semaine {w}: pairs N=1/jour, admin={per_day[JOURS[0]]['A']}, répartition J/E OK.")
    return weeks

def night_share_weights(P, cohorts, mode):
    if mode == "off":
        return {cid:1.0 for cid in cohorts}
    if mode == "global":
        total = sum(len(m) for m in cohorts.values()) or 1
        return {cid:(len(members)/total) for cid,members in cohorts.items()}
    tot_wd = sum(p["wd"] for p in P) or 1
    w = {}
    for cid, members in cohorts.items():
        w[cid] = (sum(next(p["wd"] for p in P if p["n"]==n) for n in members) / tot_wd)
    return w

# ------------------------------
# Core scheduler
# ------------------------------

def schedule(P, W, staff, edo_plan, edo_fixed,
             hard_mode, overlast,
             night_mode, night_mode_kind, eve_mode,
             share_mode, fairness_cohorts, fairness_of,
             verbose=0):
    rows = []
    names = [p["n"] for p in P]
    info = {
        p["n"]:{ 
            "cnt":{s:0 for s in ["D","E","N","A"]},
            "last":{}, "wkdays":{w:0 for w in week_range(W)},
            "nights":0, "tot":0
        } for p in P
    }
    caps    = {p["n"]: p["mxn"] for p in P}
    prefers = {p["n"]: p["pn"] for p in P}
    noE     = {p["n"]: p["ne"] for p in P}
    wd      = {p["n"]: p["wd"] for p in P}
    ttarget = {p["n"]: sum((wd[p["n"]] - (1 if p["n"] in edo_plan[w] else 0)) for w in week_range(W)) for p in P}
    cap_total = (hard_mode in ("warn","enforce"))
    weights = night_share_weights(P, fairness_cohorts, share_mode)
    quotaN = {w:{} for w in week_range(W)}
    usedN  = {w:{cid:0 for cid in fairness_cohorts} for w in week_range(W)}
    for w in week_range(W):
        total_n_slots = sum(staff[w][d]["N"]*2 for d in JOURS)
        for cid in fairness_cohorts:
            quotaN[w][cid] = weights.get(cid,1.0)*total_n_slots

    def ok_after_night(n, w, idx):
        y = JOURS[idx-1] if idx>0 else None
        return not (y and info[n]["last"].get((w,y))=="N")
    def bad_e_to_d(n, w, idx, sh):
        y = JOURS[idx-1] if idx>0 else None
        return (sh=="D" and y and info[n]["last"].get((w,y))=="E")
    def remain_total(n):
        return ttarget[n] - info[n]["tot"]
    def avg_deficit(n, shift, mode):
        if mode=="off": return 0.0
        if mode=="global":
            allv = [info[x]["cnt"][shift] for x in names]
            avg = (sum(allv)/len(allv)) if allv else 0.0
            return info[n]["cnt"][shift]-avg
        cid = fairness_of[n]
        vals = [info[x]["cnt"][shift] for x in fairness_cohorts[cid]]
        avg = (sum(vals)/len(vals)) if vals else 0.0
        return info[n]["cnt"][shift]-avg
    def rate_deficit_night(n, mode):
        rn = info[n]["cnt"]["N"] / max(1, ttarget[n])
        if mode=="off": return 0.0
        if mode=="global":
            rates = [info[x]["cnt"]["N"]/max(1,ttarget[x]) for x in names]
            avg = (sum(rates)/len(rates)) if rates else 0.0
            return rn-avg
        cid = fairness_of[n]; grp = fairness_cohorts[cid]
        rates = [info[x]["cnt"]["N"]/max(1,ttarget[x]) for x in grp]
        avg = (sum(rates)/len(rates)) if rates else 0.0
        return rn-avg
    def team_share_gap(n, w):
        cid = fairness_of[n]
        return usedN[w][cid]-quotaN[w][cid]

    def pick_candidates(w, di, day, s, need):
        fixed_off = {n for n,f in edo_fixed.items() if f==day and n in edo_plan[w]}
        allow_total_override = (overlast and hard_mode=="enforce" and w==W)
        def base_filter(n):
            return (n not in fixed_off and not info[n]["last"].get((w,day)) and ok_after_night(n,w,di))
        levels = []
        levels.append([n for n in names if base_filter(n) and (info[n]["wkdays"][w] < (wd[n] - (1 if n in edo_plan[w] else 0))) and ((not cap_total) or (remain_total(n)>0) or allow_total_override) and (s!="N" or info[n]["nights"]<caps[n])])
        levels.append([n for n in names if base_filter(n) and ((not cap_total) or (remain_total(n)>0) or allow_total_override) and (s!="N" or info[n]["nights"]<caps[n])])
        if hard_mode != "enforce" or allow_total_override:
            levels.append([n for n in names if base_filter(n) and (s!="N" or info[n]["nights"]<caps[n])])
        if s=="N" and hard_mode!="enforce":
            levels.append([n for n in names if base_filter(n)])
        for cand in levels:
            if cand:
                random.shuffle(cand)
                cand.sort(key=lambda n: (
                    (team_share_gap(n,w) if (s=="N" and share_mode!="off") else 0.0),
                    (rate_deficit_night(n, night_mode) if (s=="N" and night_mode_kind=="rate") else (avg_deficit(n,"N",night_mode) if s=="N" else 0.0)),
                    bad_e_to_d(n,w,di,s),
                    (avg_deficit(n,"E",eve_mode) if s=="E" else 0.0),
                    (s=="E" and noE[n]),
                    -(s=="N" and prefers[n]),
                    -remain_total(n),
                    info[n]["cnt"][s],
                    info[n]["wkdays"][w],
                ))
                take = cand[:need]
                if len(take) < need: continue
                return take
        return []

    for w in week_range(W):
        for di,day in enumerate(JOURS):
            for s in ["N","E","D","A"]:
                slots = staff[w][day][s]
                for ps in range(1, slots+1):
                    need = 2 if s!="A" else 1
                    take = pick_candidates(w,di,day,s,need)
                    if len(take) < need and hard_mode=="enforce" and not (overlast and w==W):
                        left = [n for n in names if (not info[n]["last"].get((w,day)) and ok_after_night(n,w,di))]
                        random.shuffle(left); take = (take + left)[:need]
                    row = [w,day,s,ps]
                    if need==2:
                        a,b = (take+["",""])[:2]
                        row += [a,b]
                        for n in [a,b]:
                            if n:
                                info[n]["cnt"][s]+=1; info[n]["wkdays"][w]+=1; info[n]["last"][(w,day)]=s; info[n]["tot"]+=1
                                if s=="N":
                                    info[n]["nights"]+=1
                    else:
                        a = (take+[""])[0]; row += [a,""]
                        if a:
                            info[a]["cnt"][s]+=1; info[a]["wkdays"][w]+=1; info[a]["last"][(w,day)]=s; info[a]["tot"]+=1
                    rows.append(row)

    usedN = {w:{} for w in week_range(W)}; quotaN = {w:{} for w in week_range(W)}
    return rows, info, ttarget, usedN, quotaN

# ------------------------------
# Validation + fairness scoring
# ------------------------------

def validate(rows, info, W, P, edo_plan, ttarget, fairness_cohorts):
    dup = 0; vacancies = 0
    for w in week_range(W):
        for d in JOURS:
            seen = set()
            for r in rows:
                if r[0]==w and r[1]==d:
                    a,b = r[4], r[5]
                    if not a or (r[2]!="A" and not b):
                        vacancies += 1
                    for n in [a,b]:
                        if n:
                            if n in seen: dup += 1
                            seen.add(n)
    n2o=e2d=0
    for w in week_range(W):
        for i,d in enumerate(JOURS):
            if i==0: continue
            y = JOURS[i-1]
            for n in info:
                if info[n]["last"].get((w,y))=="N" and info[n]["last"].get((w,d)):
                    n2o += 1
                if info[n]["last"].get((w,y))=="E" and info[n]["last"].get((w,d))=="D":
                    e2d += 1
    wd = {p["n"]:p["wd"] for p in P}
    mis=0
    for n in info:
        for w in week_range(W):
            cible = wd[n] - (1 if n in edo_plan[w] else 0)
            if info[n]["wkdays"][w] != cible:
                mis += 1
    tot_mis = sum(1 for n in info if info[n]["tot"] != ttarget[n])
    return {
        "doublons_jour":dup,
        "Nuit_suivie_travail":n2o,
        "Soir_vers_Jour":e2d,
        "Ecarts_hebdo_jours":mis,
        "Ecarts_horizon_personnes":tot_mis,
        "Slots_vides":vacancies,
    }

def fairness_std(info, fairness_cohorts, shift):
    if not fairness_cohorts:
        return 0.0
    total = 0.0
    for cid, members in fairness_cohorts.items():
        vals = [info[n]["cnt"][shift] for n in members] if members else [0]
        total += pstdev(vals) if len(vals) > 1 else 0.0
    return total

def score_solution(vals, info, fairness_cohorts):
    stdN = fairness_std(info, fairness_cohorts, "N")
    stdE = fairness_std(info, fairness_cohorts, "E")
    score = (
        10*vals.get("Slots_vides",0) +
        5*vals.get("doublons_jour",0) +
        3*vals.get("Nuit_suivie_travail",0) +
        1*vals.get("Soir_vers_Jour",0) +
        2*vals.get("Ecarts_hebdo_jours",0) +
        2*vals.get("Ecarts_horizon_personnes",0) +
        10*stdN + 3*stdE
    )
    return float(score), {"stdN":stdN, "stdE":stdE}

# ------------------------------
# Excel helpers
# ------------------------------

def write_week_headers(ws, W, start_col=2, row=1):
    c = start_col
    for w in week_range(W):
        ws.merge_cells(start_row=row, start_column=c, end_row=row, end_column=c+4)
        ws.cell(row=row, column=c, value=f"SEMAINE {w}").font = Font(bold=True)
        ws.cell(row=row, column=c).alignment = Alignment(horizontal="center", vertical="center")
        c += 5

def write_days_row(ws, W, start_col=2, row=2):
    c = start_col
    for w in week_range(W):
        for d in JOURS:
            cell = ws.cell(row=row, column=c, value=d)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            c += 1

def write_col_widths(ws, total_cols, width=14):
    for i in range(1, total_cols+1):
        ws.column_dimensions[get_column_letter(i)].width = width

def _merge_border(existing, side, new_side):
    return Border(
        left  = new_side if side=="left"  else existing.left,
        right = new_side if side=="right" else existing.right,
        top   = new_side if side=="top"   else existing.top,
        bottom= new_side if side=="bottom"else existing.bottom,
        diagonal=existing.diagonal, outline=existing.outline, vertical=existing.vertical, horizontal=existing.horizontal
    )

def apply_week_separators(ws, W, start_col):
    max_row = ws.max_row
    for w in week_range(W):
        cstart = start_col + (w-1)*5
        cend   = cstart + 4
        for r in range(1, max_row+1):
            ws.cell(row=r, column=cstart).border = _merge_border(ws.cell(row=r, column=cstart).border or Border(), "left", DOUBLEK)
            ws.cell(row=r, column=cend).border   = _merge_border(ws.cell(row=r, column=cend).border or Border(),   "right", DOUBLEK)
        for c in range(cstart, cend+1):
            ws.cell(row=1, column=c).border        = _merge_border(ws.cell(row=1, column=c).border or Border(), "top", DOUBLEK)
            ws.cell(row=max_row, column=c).border  = _merge_border(ws.cell(row=max_row, column=c).border or Border(), "bottom", DOUBLEK)

def shade_week_headers(ws, W, start_col, rows=(1,2)):
    TINT = "F5F5F5"
    fill = PatternFill(start_color=TINT, end_color=TINT, fill_type="solid")
    for w in week_range(W):
        if w % 2 == 0:
            cstart = start_col + (w-1)*5
            cend   = cstart + 4
            for r in rows:
                for c in range(cstart, cend+1):
                    cell = ws.cell(row=r, column=c)
                    if not (cell.fill and cell.fill.fill_type == "solid"):
                        cell.fill = fill

def draw_rect_border(ws, r1, c1, r2, c2, side):
    for c in range(c1, c2+1):
        ws.cell(row=r1, column=c).border = _merge_border(ws.cell(row=r1, column=c).border or Border(), "top", side)
        ws.cell(row=r2, column=c).border = _merge_border(ws.cell(row=r2, column=c).border or Border(), "bottom", side)
    for r in range(r1, r2+1):
        ws.cell(row=r, column=c1).border = _merge_border(ws.cell(row=r, column=c1).border or Border(), "left", side)
        ws.cell(row=r, column=c2).border = _merge_border(ws.cell(row=r, column=c2).border or Border(), "right", side)

def build_matrice_fused(df_assign: pd.DataFrame, people, weeks: int, edo_plan: dict, info: dict) -> pd.DataFrame:
    names   = sorted(info.keys())
    headers = [f"S{w} {d}" for w in week_range(weeks) for d in JOURS]
    mat = pd.DataFrame(index=names, columns=headers).fillna("")
    for _,r in df_assign.iterrows():
        c = f"S{int(r.Semaine)} {r.Jour}"; s = r.Poste
        if r.Pers_A: mat.at[r.Pers_A, c] = CODE[s]
        if r.Pers_B: mat.at[r.Pers_B, c] = CODE[s]
    edof_map = {p["n"]: (p.get("edof") or "") for p in people}
    for n in names:
        for w in week_range(weeks):
            is_edo_week = (n in edo_plan[w]); fixed = edof_map.get(n, "")
            empties = []
            for d in JOURS:
                col = f"S{w} {d}"
                if not str(mat.at[n, col]).strip():
                    mat.at[n, col] = "OFF"; empties.append(col)
            if is_edo_week:
                target = f"S{w} {fixed}" if fixed in JOURS else None
                if target and mat.at[n, target] == "OFF":
                    mat.at[n, target] = "EDO"
                else:
                    if empties:
                        first_off = empties[0]
                        mat.at[n, first_off] = "EDO*" if (fixed in JOURS and target and mat.at[n, target] != "OFF") else "EDO"
    return mat

def build_counts_by_shift(df_assign, weeks):
    headers = [f"S{w} {d}" for w in week_range(weeks) for d in JOURS]
    counts = pd.DataFrame(index=["Jour","Soir","Nuit","Admin"], columns=headers).fillna(0)
    counts = counts.infer_objects(copy=False)  # <= recommandé par pandas pour éviter le downcasting implicite
    for _,r in df_assign.iterrows():
        col = f"S{int(r.Semaine)} {r.Jour}"
        if r.Poste in ("D","E","N"):
            if r.Pers_A and str(r.Pers_A).strip(): counts.at[{"D":"Jour","E":"Soir","N":"Nuit"}[r.Poste], col] += 1
            if r.Pers_B and str(r.Pers_B).strip(): counts.at[{"D":"Jour","E":"Soir","N":"Nuit"}[r.Poste], col] += 1
        elif r.Poste == "A":
            if r.Pers_A and str(r.Pers_A).strip(): counts.at["Admin", col] += 1
    return counts.astype(int)

def build_weekly_hours_map(info, W):
    m = {}
    for n in info:
        for w in week_range(W):
            h=0
            for d in JOURS:
                s = info[n]["last"].get((w,d))
                if s: h += HEURES[s]
            m[(n,w)] = h
    return m

def per_person_issues(info, W):
    iss = {n:{"soir_jour":0, "nuit_travail":0} for n in info}
    for n in info:
        for w in week_range(W):
            for i,d in enumerate(JOURS):
                if i==0: continue
                prev = JOURS[i-1]
                s_prev = info[n]["last"].get((w,prev))
                s_cur  = info[n]["last"].get((w,d))
                if s_prev == "N" and s_cur: iss[n]["nuit_travail"] += 1
                if s_prev == "E" and s_cur == "D": iss[n]["soir_jour"] += 1
    return iss

def per_person_week_devs(info, P, edo_plan, W):
    wd = {p["n"]:p["wd"] for p in P}
    devs = {n:[] for n in info}
    for n in info:
        for w in week_range(W):
            cible = wd[n] - (1 if n in edo_plan[w] else 0)
            worked = info[n]["wkdays"][w]
            if worked != cible:
                devs[n].append((w, worked-cible))
    return devs

def team_label_of(p):
    if p.get("team"):
        return p["team"]
    return f"Équipe {p['wd']}j"

def pct_from_wd(wd):
    return f"{int(round(100*wd/5))}%"

def exports(rows, info, P, W, xlsx_path, csv_path, edo_plan, vals, opts, ttarget, fairness_cohorts, usedN, quotaN, team_borders=False):
    cols = ["Semaine","Jour","Poste","Pair","Pers_A","Pers_B"]
    df = pd.DataFrame(rows, columns=cols)
    if csv_path: df.to_csv(csv_path, index=False)

    mat = build_matrice_fused(df, P, W, edo_plan, info)
    counts = build_counts_by_shift(df, W)

    by = {}
    headers = [f"S{w} {d}" for w in week_range(W) for d in JOURS]
    for w in week_range(W):
        for d in JOURS:
            for s in ["D","E","N","A"]:
                x = df[(df.Semaine==w)&(df.Jour==d)&(df.Poste==s)]
                pairs = []
                for _,r in x.iterrows():
                    if s=="A": pairs.append(r.Pers_A or "")
                    else: pairs.append(" / ".join([r.Pers_A or "", r.Pers_B or ""]))
                by[(s, f"S{w} {d}")] = "; ".join([p for p in pairs if p.strip()])
    poste = pd.DataFrame(index=["Jour","Soir","Nuit","Admin"], columns=headers)
    mname = {"D":"Jour","E":"Soir","N":"Nuit","A":"Admin"}
    for (s,c),v in by.items():
        poste.at[mname[s], c] = v

    tot = []
    for p in P:
        n = p["n"]; c = info[n]["cnt"]; worked_total = info[n]["tot"]
        hrs = c["D"]*10 + c["E"]*10 + c["N"]*12 + c["A"]*8
        ttarget_n = ttarget[n]
        nb_edo = sum(1 for w in week_range(W) if n in edo_plan[w])
        tot.append([n, c["D"], c["E"], c["N"], c["A"], worked_total, hrs, ttarget_n, worked_total-ttarget_n, p["wd"], nb_edo])
    syn = pd.DataFrame(tot, columns=["Nom","Jours","Soirs","Nuits","Admin","Total_Jours","Heures","Cible_Totale_Jours","Ecart_Total_Jours","WD_per_week","Nb_EDO"])

    vrows = []
    wd = {p["n"]:p["wd"] for p in P}
    for p in P:
        n = p["n"]
        for w in week_range(W):
            cible  = wd[n] - (1 if n in edo_plan[w] else 0)
            worked = info[n]["wkdays"][w]
            vrows.append([n, w, "EDO" if (n in edo_plan[w]) else "", cible, worked, worked-cible])
    valid = pd.DataFrame(vrows, columns=["Nom","Semaine","EDO","Cible_Jours","Jours_Travaillés","Ecart"])

    weekly_hours = build_weekly_hours_map(info, W)
    conf = []
    for p in P:
        n = p["n"]
        for w in week_range(W):
            h = weekly_hours[(n,w)]
            conf.append([n,w,h,1 if h>48 else 0])
    c48 = pd.DataFrame(conf, columns=["Nom","Semaine","Heures",">48h"])

    fair_rows = []
    fairness_cohorts = fairness_cohorts or {"all": sorted([p["n"] for p in P])}
    for cid, members in fairness_cohorts.items():
        if not members: continue
        avgD = sum(info[n]["cnt"]["D"] for n in members)/len(members)
        avgE = sum(info[n]["cnt"]["E"] for n in members)/len(members)
        avgN = sum(info[n]["cnt"]["N"] for n in members)/len(members)
        fair_rows.append([cid, round(avgD,2), round(avgE,2), round(avgN,2)])
    inter = pd.DataFrame(fair_rows, columns=["Cohorte","Moy. Jours","Moy. Soirs","Moy. Nuits"])

    dash = pd.DataFrame({
        "Indicateur":[
            "Effectif","Semaines","Tous les postes pourvus",
            "Violations Nuit→Travail","Transitions Soir→Jour",
            "Écarts hebdo (nb de cases)","Totaux horizon non conformes (nb pers)"
        ],
        "Valeur":[
            len(mat.index), W,
            "Oui" if (sum(1 for r in rows if (r[2]!='A' and (not r[4] or not r[5])) or (r[2]=='A' and not r[4]))==0) else "Non",
            sum(1 for w in week_range(W) for i in range(1,5) for n in info if info[n]["last"].get((w,JOURS[i-1]))=="N" and info[n]["last"].get((w,JOURS[i]))),
            sum(1 for w in week_range(W) for i in range(1,5) for n in info if info[n]["last"].get((w,JOURS[i-1]))=="E" and info[n]["last"].get((w,JOURS[i]))=="D"),
            sum(1 for n in info for w in week_range(W) if info[n]["wkdays"][w] != (wd[n] - (1 if n in edo_plan[w] else 0))),
            sum(1 for n in info if info[n]["tot"] != sum((wd[n] - (1 if n in edo_plan[w] else 0)) for w in week_range(W))),
        ]
    })
    opts_rows = [
        ["EDO activé", "Oui" if opts.get("edo") else "Non"],
        ["Jour EDO global", opts.get("edo_fixed") or "—"],
        ["Cohortes d'équité", opts.get("fairness_cohorts")],
        ["Partage Nuits inter-cohortes", opts.get("share_mode")],
        ["Équité Nuits (portée)", opts.get("night_mode")],
        ["Équité Nuits (mode)", opts.get("night_mode_kind")],
        ["Équité Soirs (portée)", opts.get("eve_mode")],
        ["Moyenne horizon (hard mode)", opts.get("hard_mode")],
        ["Remplir dernière semaine", "Oui" if opts.get("overlast") else "Non"],
        ["Rebalance (étapes)", opts.get("post_steps")],
        ["Seed gagnante", opts.get("best_seed")],
        ["Essais (tries)", opts.get("tries")],
        ["Score (min)", opts.get("best_score")],
        ["Bordures équipes Matrice", "Oui" if team_borders else "Non"],
    ]
    runopts = pd.DataFrame(opts_rows, columns=["Option","Valeur"])
    summary_msgs = pd.DataFrame({"Résumé":[
        "✅ Tous les postes pourvus." if dash.iloc[0,1]=="Oui" else "⚠ Des postes non pourvus (voir Technique).",
        f"Seed gagnante: {opts.get('best_seed')} sur {opts.get('tries')} essais (score {opts.get('best_score')}).",
    ]})

    wb = Workbook()

    ws_db = wb.active; ws_db.title = "Tableau de bord"
    for j, col in enumerate(dash.columns, start=1):
        ws_db.cell(row=1, column=j, value=col).font = Font(bold=True)
    for i in range(len(dash)):
        for j in range(len(dash.columns)):
            ws_db.cell(row=2+i, column=1+j, value=dash.iat[i,j])
    ro_start = len(dash)+4
    for j, col in enumerate(runopts.columns, start=1):
        ws_db.cell(row=ro_start, column=j, value=col).font = Font(bold=True)
    for i in range(len(runopts)):
        for j in range(len(runopts.columns)):
            ws_db.cell(row=ro_start+1+i, column=1+j, value=runopts.iat[i,j])
    sm_start = ro_start + len(runopts) + 3
    ws_db.cell(row=sm_start-1, column=1, value="Résumé des points d’attention").font = Font(bold=True)
    for j, col in enumerate(summary_msgs.columns, start=1):
        ws_db.cell(row=sm_start, column=j, value=col).font = Font(bold=True)
    for i in range(len(summary_msgs)):
        ws_db.cell(row=sm_start+1+i, column=1, value=summary_msgs.iloc[i,0])
    for i in range(1, 12):
        ws_db.column_dimensions[get_column_letter(i)].width = 24
    ws_db.freeze_panes = "A2"

    # -------- Matrice (with aligned summary & optional team borders)
    ws_m = wb.create_sheet("Matrice")
    write_week_headers(ws_m, W, start_col=2, row=1)
    write_days_row(ws_m, W, start_col=2, row=2)
    ws_m.freeze_panes = "B3"

    groups = {}
    for p in P:
        key = p.get("team") or f"{p['wd']}j"
        groups.setdefault(key, []).append(p["n"])
    for k in groups:
        groups[k] = sorted(groups[k])
    ordered_people = []
    for key in sorted(groups.keys(), key=lambda x: (len(groups[x]), x), reverse=True):
        ordered_people.extend(groups[key])

    row_map = {}
    for r, n in enumerate(ordered_people, start=3):
        row_map[n] = r
        ws_m.cell(row=r, column=1, value=n)
        c=2
        for w in week_range(W):
            for d in JOURS:
                val = str(mat.at[n, f"S{w} {d}"])
                cell = ws_m.cell(row=r, column=c, value=val)
                if val in COLORS:
                    cell.fill = PatternFill(start_color=COLORS[val], end_color=COLORS[val], fill_type="solid")
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = BORDER_THIN
                c+=1
    write_col_widths(ws_m, ws_m.max_column, 14)

    summary_start = ws_m.max_row + 2
    ws_m.cell(row=summary_start-1, column=1, value="Résumé — # personnes par poste").font = Font(bold=True)
    summary_rows = {"Jour": summary_start, "Soir": summary_start+1, "Nuit": summary_start+2, "Admin": summary_start+3}
    for label, rr in summary_rows.items():
        ws_m.cell(row=rr, column=1, value=label).font = Font(bold=True)
        c=2
        for w in week_range(W):
            for d in JOURS:
                val = int(counts.at[label, f"S{w} {d}"]) if f"S{w} {d}" in counts.columns else 0
                cell = ws_m.cell(row=rr, column=c, value=val)
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = BORDER_THIN
                tint = COLORS[{"Jour":"J","Soir":"S","Nuit":"N","Admin":"A"}[label]]
                cell.fill = PatternFill(start_color=tint, end_color=tint, fill_type="solid")
                c+=1

    shade_week_headers(ws_m, W, start_col=2, rows=(1,2))
    apply_week_separators(ws_m, W, start_col=2)

    if team_borders:
        total_cols = 1 + 5*W
        blocks = []
        i=0; N=len(ordered_people)
        while i<N:
            p = ordered_people[i]
            team_key = next((pp.get("team") or f"{pp['wd']}j" for pp in P if pp["n"]==p), "group")
            j=i
            while j+1<N:
                pn = ordered_people[j+1]
                team2 = next((pp.get("team") or f"{pp['wd']}j" for pp in P if pp["n"]==pn), "group")
                if team2!=team_key: break
                j+=1
            r1 = row_map[ordered_people[i]]
            r2 = row_map[ordered_people[j]]
            blocks.append((r1, r2, team_key))
            i = j+1
        for (r1,r2,_key) in blocks:
            draw_rect_border(ws_m, r1, 1, r2, total_cols, BLACKTHIN)

    # -------- ParPoste_Statique
    ws_pp = wb.create_sheet("ParPoste_Statique")
    write_week_headers(ws_pp, W, start_col=2, row=1)
    write_days_row(ws_pp, W, start_col=2, row=2)
    ws_pp.freeze_panes = "B3"
    shifts_order = ["Jour","Soir","Nuit","Admin"]
    for i, row_name in enumerate(shifts_order, start=3):
        ws_pp.cell(row=i, column=1, value=row_name).font = Font(bold=True if row_name!="Admin" else False)
        c=2
        for w in week_range(W):
            for d in JOURS:
                v = str(poste.at[row_name, f"S{w} {d}"]) if f"S{w} {d}" in poste.columns else ""
                cell = ws_pp.cell(row=i, column=c, value=v if v!="nan" else "")
                cell.alignment = Alignment(horizontal="left", vertical="center")
                cell.border = BORDER_THIN
                c+=1
    write_col_widths(ws_pp, ws_pp.max_column, 18)
    shade_week_headers(ws_pp, W, start_col=2, rows=(1,2))
    apply_week_separators(ws_pp, W, start_col=2)

    # -------- Synthèse
    ws_syn = wb.create_sheet("Synthèse")
    for j, col in enumerate(list(syn.columns), start=1):
        ws_syn.cell(row=1, column=j, value=col).font = Font(bold=True)
    for i in range(len(syn)):
        for j in range(len(syn.columns)):
            ws_syn.cell(row=2+i, column=1+j, value=syn.iat[i,j])
    for i in range(1, ws_syn.max_column+1):
        ws_syn.column_dimensions[get_column_letter(i)].width = 16
    ws_syn.freeze_panes = "A2"

    # -------- Technique
    ws_tech = wb.create_sheet("Technique")
    for j, col in enumerate(list(valid.columns), start=1):
        ws_tech.cell(row=1, column=j, value=col).font = Font(bold=True)
    for i in range(len(valid)):
        for j in range(len(valid.columns)):
            ws_tech.cell(row=2+i, column=1+j, value=valid.iat[i,j])
    start48 = len(valid)+4
    ws_tech.cell(row=start48-1, column=1, value="Conformité 48h").font = Font(bold=True)
    for j, col in enumerate(list(c48.columns), start=1):
        ws_tech.cell(row=start48, column=j, value=col).font = Font(bold=True)
    for i in range(len(c48)):
        for j in range(len(c48.columns)):
            ws_tech.cell(row=start48+1+i, column=1+j, value=c48.iat[i,j])
    starteq = start48 + len(c48) + 4
    ws_tech.cell(row=starteq-1, column=1, value="Équité inter-cohortes").font = Font(bold=True)
    for j, col in enumerate(list(inter.columns), start=1):
        ws_tech.cell(row=starteq, column=j, value=col).font = Font(bold=True)
    for i in range(len(inter)):
        for j in range(len(inter.columns)):
            ws_tech.cell(row=starteq+1+i, column=1+j, value=inter.iat[i,j])
    for i in range(1, 12):
        ws_tech.column_dimensions[get_column_letter(i)].width = 18
    try:
        headers = [ws_tech.cell(row=1, column=j).value for j in range(1, ws_tech.max_column+1)]
        if "Ecart" in headers:
            col_idx = headers.index("Ecart")+1
            col_letter = get_column_letter(col_idx)
            ws_tech.conditional_formatting.add(f"{col_letter}2:{col_letter}{1+len(valid)}",
                CellIsRule(operator="notEqual", formula=["0"], fill=PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")))
    except:
        pass

    # -------- Per-person sheets (header + identity panel + table)
    issues = per_person_issues(info, W)
    devs   = per_person_week_devs(info, P, edo_plan, W)
    weekly_hours = build_weekly_hours_map(info, W)
    mat_for_pp = mat

    for p in P:
        n = p["n"]
        sname = ("Planning — " + n)[:31]
        ws = wb.create_sheet(title=sname)

        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=8)
        tcell = ws.cell(row=1, column=1, value=n)
        tcell.font = Font(bold=True, size=16)
        tcell.alignment = Alignment(horizontal="left", vertical="center")

        team   = team_label_of(p)
        pct    = pct_from_wd(p["wd"])
        prefs  = []
        if p["pn"]: prefs.append("préfère Nuits")
        if p["ne"]: prefs.append("éviter Soirs")
        pref_s = ", ".join(prefs) if prefs else "—"
        mxn    = p["mxn"] if p["mxn"]<10**6 else "—"
        edo_elig = "Oui" if p["edo"] else "Non"
        nb_edo = sum(1 for w in week_range(W) if n in edo_plan[w])
        edofix = p["edof"] if p["edof"] else "—"

        ws.cell(row=2, column=1, value="Équipe").font = Font(bold=True)
        ws.cell(row=2, column=2, value=team)
        ws.cell(row=2, column=3, value="% temps").font = Font(bold=True)
        ws.cell(row=2, column=4, value=pct)
        ws.cell(row=2, column=5, value="WD/sem.").font = Font(bold=True)
        ws.cell(row=2, column=6, value=p["wd"])

        ws.cell(row=3, column=1, value="Préférences").font = Font(bold=True)
        ws.cell(row=3, column=2, value=pref_s)
        ws.cell(row=3, column=3, value="Max Nuits").font = Font(bold=True)
        ws.cell(row=3, column=4, value=mxn)
        ws.cell(row=3, column=5, value="EDO éligible").font = Font(bold=True)
        ws.cell(row=3, column=6, value=edo_elig)
        ws.cell(row=3, column=7, value="Nb EDO").font = Font(bold=True)
        ws.cell(row=3, column=8, value=nb_edo)
        ws.cell(row=3, column=9, value="EDO fixe").font = Font(bold=True)
        ws.cell(row=3, column=10, value=edofix)

        start_row = 5
        headers_pp = ["Semaine"] + JOURS + ["Total Semaine","Heures"]
        for j,h in enumerate(headers_pp, start=1):
            ws.cell(row=start_row, column=j, value=h).font = Font(bold=True)
            ws.column_dimensions[get_column_letter(j)].width = 16 if j==1 else 14
        ws.freeze_panes = f"A{start_row+1}"

        for i,w in enumerate(week_range(W), start=start_row+1):
            ws.cell(row=i, column=1, value=f"S{w}")
            week_total=0; week_hours=0
            for j,d in enumerate(JOURS, start=2):
                disp = str(mat_for_pp.at[n, f"S{w} {d}"])
                cell = ws.cell(row=i, column=j, value=disp)
                if disp in COLORS:
                    cell.fill = PatternFill(start_color=COLORS[disp], end_color=COLORS[disp], fill_type="solid")
                cell.alignment = Alignment(horizontal="center", vertical="center"); cell.border = BORDER_THIN
                if disp in ("J","S","N","A"):
                    week_total += 1; week_hours += HEURES[REVCODE[disp]]
            ws.cell(row=i, column=len(JOURS)+2, value=week_total).border = BORDER_THIN
            ws.cell(row=i, column=len(JOURS)+3, value=week_hours).border = BORDER_THIN
            ws.cell(row=i, column=len(JOURS)+2).alignment = Alignment(horizontal="center")
            ws.cell(row=i, column=len(JOURS)+3).alignment = Alignment(horizontal="center")

    wb.save(xlsx_path)

# ------------------------------
# Post-pass greedy rebalance
# ------------------------------

def post_rebalance(rows, info, W, caps, ttarget, fairness_of, share_mode, usedN, quotaN, max_steps=200, verbose=0):
    day_assign = {}
    for r in rows:
        w,d,s,_,a,b = r
        day_assign.setdefault((w,d), set())
        for n in [a,b]:
            if n: day_assign[(w,d)].add(n)
    def dev(n): return ttarget[n]-info[n]["tot"]
    steps=0; changed=True
    while changed and steps < max_steps:
        changed=False; steps += 1
        if verbose>=1 and steps % 20 == 1:
            print(f"{C_DIM}   ↳ Rebalance: itération {steps}/{max_steps}{C_RESET}")
        for r in rows:
            w,d,s,_,a,b = r
            need = 1 if s=="A" else 2
            assigned = [a] if need==1 else [a,b]
            for pos,x in enumerate(assigned):
                if not x or dev(x) >= 0: continue
                for y in list(info.keys()):
                    if y in day_assign[(w,d)]: continue
                    i = JOURS.index(d)
                    if i>0 and info[y]["last"].get((w,JOURS[i-1]))=="N": continue
                    if s=="N" and info[y]["nights"] >= caps[y]: continue
                    if dev(y) <= 0: continue
                    assigned[pos] = y
                    if need==1: r[4]=y
                    else: r[4 if pos==0 else 5] = y
                    day_assign[(w,d)].discard(x); day_assign[(w,d)].add(y)
                    info[x]["cnt"][s]-=1; info[x]["wkdays"][w]-=1; info[x]["tot"]-=1
                    info[y]["cnt"][s]+=1; info[y]["wkdays"][w]+=1; info[y]["tot"]+=1
                    if s=="N": info[x]["nights"]-=1; info[y]["nights"]+=1
                    changed=True; break
                if changed: break
            if changed: break
    return rows, info

# ------------------------------
# CLI entry
# ------------------------------

def main():
    ap = argparse.ArgumentParser(description="Rota Optimizer — v2.9 (matrice summary + team borders + person headers)")
    ap.add_argument("--pattern-csv", required=True)
    ap.add_argument("--weeks", type=int, default=12)
    ap.add_argument("--xlsx-path", default="rota.xlsx")
    ap.add_argument("--csv-path", default="rota.csv")
    ap.add_argument("--edo", action="store_true")
    ap.add_argument("--edo-jour-fixe", choices=JOURS)
    ap.add_argument("--fairness-cohorts", choices=["none","by-wd"], default="by-wd")
    ap.add_argument("--hard-average-mode", choices=["off","warn","enforce"], default="warn")
    ap.add_argument("--overassign-last-week", action="store_true")
    ap.add_argument("--night-fairness", choices=["off","global","cohort"], default="cohort")
    ap.add_argument("--night-fairness-mode", choices=["count","rate"], default="rate")
    ap.add_argument("--evening-fairness", choices=["off","global","cohort"], default="cohort")
    ap.add_argument("--inter-team-night-share", choices=["off","proportional","global"], default="proportional")
    ap.add_argument("--post-rebalance-steps", type=int, default=300)
    ap.add_argument("--no-spinner", action="store_true", help="Désactiver les animations CLI")
    ap.add_argument("-v", "--verbose", action="count", default=0, help="-v (semaine), -vv (jour/slot)")
    ap.add_argument("--seed", type=int, help="Graine RNG (reproductible). Si utilisée avec --tries, on teste seed, seed+1, ...")
    ap.add_argument("--tries", type=int, default=1, help="Nombre d'essais; on garde le meilleur")
    ap.add_argument("--matrice-team-borders", action="store_true", help="Encadre finement les blocs d'équipes (détectées par 'team' ou WD/sem).")
    a = ap.parse_args()

    with Spinner("Chargement du CSV", enabled=not a.no_spinner):
        P = load_people(a.__dict__["pattern_csv"])
        W = a.weeks

    with Spinner("Préparation des cohortes & EDO", enabled=not a.no_spinner):
        fairness_cohorts, fairness_of = make_cohorts_fairness(P, a.fairness_cohorts)
        if a.edo:
            edo_plan, fixed = build_edo_plan(P, W, a.__dict__["edo_jour_fixe"])
        else:
            edo_plan = {w:set() for w in week_range(W)}
            fixed    = {n:"" for n in [p["n"] for p in P]}

    with Spinner("Calcul des besoins par jour/poste", enabled=not a.no_spinner):
        staff = derive_staffing(P, W, edo_plan, verbose=a.verbose)

    best = None
    best_score = float("inf")
    best_seed = None
    base_seed = a.seed if a.seed is not None else int(time.time())
    tries = max(1, a.tries)

    with Spinner(f"Affectation sur {tries} essai(s) (solveur glouton)", enabled=not a.no_spinner):
        for t in range(tries):
            cur_seed = base_seed + t
            random.seed(cur_seed)
            rows, info, ttarget, usedN, quotaN = schedule(
                P, W, staff, edo_plan, fixed,
                a.hard_average_mode, a.overassign_last_week,
                a.night_fairness, a.night_fairness_mode, a.evening_fairness,
                a.inter_team_night_share, fairness_cohorts, fairness_of,
                verbose=(a.verbose if tries==1 else 0)
            )
            if a.post_rebalance_steps > 0:
                caps = {p["n"]:p["mxn"] for p in P}
                rows, info = post_rebalance(
                    rows, info, W, caps, ttarget, fairness_of, a.inter_team_night_share, usedN, quotaN,
                    max_steps=a.post_rebalance_steps, verbose=(a.verbose if tries==1 else 0)
                )
            vals = validate(rows, info, W, P, edo_plan, ttarget, fairness_cohorts)
            score, extras = score_solution(vals, info, fairness_cohorts)
            if score < best_score:
                best_score = score; best_seed = cur_seed
                best = (rows, info, ttarget, usedN, quotaN, vals)

    rows, info, ttarget, usedN, quotaN, vals = best

    with Spinner(f"Export Excel/CSV → {a.xlsx_path}", enabled=not a.no_spinner):
        exports(
            rows, info, P, W, a.xlsx_path, a.csv_path, edo_plan, vals,
            {
                "weeks":W, "edo":a.edo, "edo_fixed":a.__dict__["edo_jour_fixe"],
                "fairness_cohorts":a.fairness_cohorts, "hard_mode":a.hard_average_mode,
                "overlast":a.overassign_last_week, "night_mode":a.night_fairness,
                "night_mode_kind":a.night_fairness_mode, "eve_mode":a.evening_fairness,
                "share_mode":a.inter_team_night_share, "post_steps":a.post_rebalance_steps,
                "best_seed":best_seed, "tries":tries, "best_score":round(best_score,2)
            },
            ttarget, fairness_cohorts, usedN, quotaN, team_borders=a.matrice_team_borders
        )

    k_all_filled = (vals.get("Slots_vides",0) == 0)
    ok_badge = f"{C_OK}OK{C_RESET}" if k_all_filled else f"{C_WARN}ATTENTION{C_RESET}"
    lines = [
        f"{C_BOLD}Effectif:{C_RESET} {len(P)}   {C_BOLD}Semaines:{C_RESET} {W}   {C_BOLD}Fichier:{C_RESET} {a.xlsx_path}",
        f"{C_BOLD}Seed gagnante:{C_RESET} {best_seed}   {C_BOLD}Essais:{C_RESET} {tries}   {C_BOLD}Score:{C_RESET} {best_score:.2f}",
        f"{C_BOLD}Tous les postes pourvus:{C_RESET} {'Oui' if k_all_filled else 'Non'}",
        f"{C_BOLD}Nuit→Travail:{C_RESET} {vals.get('Nuit_suivie_travail',0)}   {C_BOLD}Soir→Jour:{C_RESET} {vals.get('Soir_vers_Jour',0)}",
        f"{C_BOLD}Écarts hebdo (cases):{C_RESET} {vals.get('Ecarts_hebdo_jours',0)}   {C_BOLD}Totaux horizon non conformes:{C_RESET} {vals.get('Ecarts_horizon_personnes',0)}",
    ]
    box(f"Rota Optimizer — Résultat ({ok_badge})", lines)
    print(f"{C_INFO}Terminé.{C_RESET} Excel: {a.xlsx_path}   CSV: {a.csv_path}   Seed: {best_seed}   Score: {best_score:.2f}")

if __name__ == "__main__":
    main()
