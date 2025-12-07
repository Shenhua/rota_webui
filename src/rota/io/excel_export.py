"""Excel export functionality for schedules."""
from pathlib import Path
from typing import List, Dict, Any, Optional, Union
import io

import pandas as pd
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.formatting.rule import CellIsRule

from rota.models.person import Person
from rota.models.schedule import Schedule
from rota.models.shift import ShiftType, WEEKDAYS, WEEKEND, ALL_DAYS


# Color scheme for shifts
SHIFT_COLORS = {
    "J": "DDEEFF",
    "S": "FFE4CC",
    "N": "E6CCFF",
    "A": "DDDDDD",
    "OFF": "EEEEEE",
    "EDO": "D8D8D8",
    "EDO*": "FFC7CE",
}

# Border styles
THIN = Side(border_style="thin", color="CCCCCC")
MEDIUM = Side(border_style="medium", color="BBBBBB")
DOUBLE_BLACK = Side(border_style="double", color="000000")
BLACK_THIN = Side(border_style="thin", color="000000")
BORDER_THIN = Border(top=THIN, bottom=THIN, left=THIN, right=THIN)


def _week_range(weeks: int) -> List[int]:
    """Generate list of week numbers."""
    return list(range(1, weeks + 1))


def _write_week_headers(ws, weeks: int, days: List[str], start_col: int = 2, row: int = 1):
    """Write week headers with merged cells."""
    c = start_col
    days_count = len(days)
    for w in _week_range(weeks):
        ws.merge_cells(start_row=row, start_column=c, end_row=row, end_column=c + days_count - 1)
        cell = ws.cell(row=row, column=c, value=f"SEMAINE {w}")
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        c += days_count


def _write_days_row(ws, weeks: int, days: List[str], start_col: int = 2, row: int = 2):
    """Write day headers for each week."""
    c = start_col
    for _ in _week_range(weeks):
        for d in days:
            cell = ws.cell(row=row, column=c, value=d)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            c += 1


def _apply_week_separators(ws, weeks: int, days: List[str], start_col: int = 2):
    """Add double borders between weeks."""
    max_row = ws.max_row
    days_count = len(days)
    for w in _week_range(weeks):
        cstart = start_col + (w - 1) * days_count
        cend = cstart + days_count - 1
        for r in range(1, max_row + 1):
            cell_left = ws.cell(row=r, column=cstart)
            cell_right = ws.cell(row=r, column=cend)
            # Merge existing border with new separator
            if cell_left.border:
                cell_left.border = Border(
                    left=DOUBLE_BLACK, right=cell_left.border.right,
                    top=cell_left.border.top, bottom=cell_left.border.bottom
                )
            if cell_right.border:
                cell_right.border = Border(
                    left=cell_right.border.left, right=DOUBLE_BLACK,
                    top=cell_right.border.top, bottom=cell_right.border.bottom
                )


def _build_matrix_df(schedule: Schedule, people: List[Person], days: List[str]) -> pd.DataFrame:
    """Build person × day matrix from schedule."""
    names = [p.name for p in people]
    weeks = schedule.weeks
    headers = [f"S{w} {d}" for w in _week_range(weeks) for d in days]
    
    mat = pd.DataFrame(index=names, columns=headers).fillna("")
    
    df = schedule.to_dataframe()
    for _, row in df.iterrows():
        w, d, shift = row["week"], row["day"], row["shift"]
        if d in days:
            col = f"S{w} {d}"
            if col in mat.columns:
                mat.at[row["name"], col] = shift
    
    # Fill empty cells with OFF
    for n in names:
        for col in headers:
            if not str(mat.at[n, col]).strip():
                mat.at[n, col] = "OFF"
    
    return mat


def _build_counts_df(schedule: Schedule, days: List[str]) -> pd.DataFrame:
    """Build shift counts per day."""
    weeks = schedule.weeks
    headers = [f"S{w} {d}" for w in _week_range(weeks) for d in days]
    counts = pd.DataFrame(0, index=["Jour", "Soir", "Nuit", "Admin"], columns=headers)
    
    df = schedule.to_dataframe()
    shift_map = {"J": "Jour", "S": "Soir", "N": "Nuit", "A": "Admin"}
    
    for _, row in df.iterrows():
        w, d, shift = row["week"], row["day"], row["shift"]
        if d in days and shift in shift_map:
            col = f"S{w} {d}"
            if col in counts.columns:
                counts.at[shift_map[shift], col] += 1
    
    return counts.astype(int)


def export_to_excel(
    schedule: Schedule,
    people: List[Person],
    output: Union[str, Path, io.BytesIO],
    days: Optional[List[str]] = None,
    team_borders: bool = False,
    config: Optional[Dict[str, Any]] = None,
) -> None:
    """
    Export schedule to Excel workbook.
    
    Args:
        schedule: Schedule object with assignments
        people: List of Person objects
        output: File path or BytesIO buffer
        days: Days to include (default: WEEKDAYS)
        team_borders: Draw borders around team groups
        config: Solver configuration for metadata
    """
    if days is None:
        days = WEEKDAYS
    
    weeks = schedule.weeks
    wb = Workbook()
    
    # Build data frames
    mat = _build_matrix_df(schedule, people, days)
    counts = _build_counts_df(schedule, days)
    
    # ========== Dashboard Sheet ==========
    ws_db = wb.active
    ws_db.title = "Tableau de bord"
    
    # Summary metrics
    summary_data = [
        ["Indicateur", "Valeur"],
        ["Effectif", len(people)],
        ["Semaines", weeks],
        ["Score", round(schedule.score, 2)],
        ["Statut", schedule.status],
        ["Temps (s)", round(schedule.solve_time_seconds, 2)],
    ]
    for i, row_data in enumerate(summary_data, start=1):
        for j, val in enumerate(row_data, start=1):
            cell = ws_db.cell(row=i, column=j, value=val)
            if i == 1:
                cell.font = Font(bold=True)
    
    # Violations
    if schedule.violations:
        start = len(summary_data) + 3
        ws_db.cell(row=start - 1, column=1, value="Violations").font = Font(bold=True)
        for i, (key, val) in enumerate(schedule.violations.items()):
            ws_db.cell(row=start + i, column=1, value=key)
            ws_db.cell(row=start + i, column=2, value=val)
    
    for i in range(1, 5):
        ws_db.column_dimensions[get_column_letter(i)].width = 24
    ws_db.freeze_panes = "A2"
    
    # ========== Matrix Sheet ==========
    ws_m = wb.create_sheet("Matrice")
    _write_week_headers(ws_m, weeks, days, start_col=2, row=1)
    _write_days_row(ws_m, weeks, days, start_col=2, row=2)
    ws_m.freeze_panes = "B3"
    
    # Group people by team
    groups: Dict[str, List[str]] = {}
    for p in people:
        key = p.team if p.team else f"{p.workdays_per_week}j"
        groups.setdefault(key, []).append(p.name)
    for k in groups:
        groups[k] = sorted(groups[k])
    
    ordered_people = []
    for key in sorted(groups.keys(), key=lambda x: (len(groups[x]), x), reverse=True):
        ordered_people.extend(groups[key])
    
    # Write matrix
    row_map = {}
    for r, name in enumerate(ordered_people, start=3):
        row_map[name] = r
        ws_m.cell(row=r, column=1, value=name)
        c = 2
        for w in _week_range(weeks):
            for d in days:
                col_key = f"S{w} {d}"
                val = str(mat.at[name, col_key]) if name in mat.index else ""
                cell = ws_m.cell(row=r, column=c, value=val)
                if val in SHIFT_COLORS:
                    fill_color = SHIFT_COLORS[val]
                    cell.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = BORDER_THIN
                c += 1
    
    # Counts summary
    summary_start = len(ordered_people) + 4
    ws_m.cell(row=summary_start, column=1, value="# personnes par poste").font = Font(bold=True)
    
    for idx, label in enumerate(["Jour", "Soir", "Nuit", "Admin"], start=1):
        row_num = summary_start + idx
        ws_m.cell(row=row_num, column=1, value=label).font = Font(bold=True)
        c = 2
        for w in _week_range(weeks):
            for d in days:
                col_key = f"S{w} {d}"
                val = int(counts.at[label, col_key]) if col_key in counts.columns else 0
                cell = ws_m.cell(row=row_num, column=c, value=val)
                cell.alignment = Alignment(horizontal="center")
                cell.border = BORDER_THIN
                shift_code = {"Jour": "J", "Soir": "S", "Nuit": "N", "Admin": "A"}[label]
                if shift_code in SHIFT_COLORS:
                    cell.fill = PatternFill(
                        start_color=SHIFT_COLORS[shift_code],
                        end_color=SHIFT_COLORS[shift_code],
                        fill_type="solid"
                    )
                c += 1
    
    _apply_week_separators(ws_m, weeks, days, start_col=2)
    
    # ========== Synthesis Sheet ==========
    ws_syn = wb.create_sheet("Synthèse")
    stats = schedule.get_person_stats()
    if not stats.empty:
        for j, col in enumerate(stats.columns, start=1):
            ws_syn.cell(row=1, column=j, value=col).font = Font(bold=True)
        for i in range(len(stats)):
            for j in range(len(stats.columns)):
                ws_syn.cell(row=2 + i, column=1 + j, value=stats.iat[i, j])
        for i in range(1, len(stats.columns) + 1):
            ws_syn.column_dimensions[get_column_letter(i)].width = 14
        ws_syn.freeze_panes = "A2"
    
    # Save
    if isinstance(output, io.BytesIO):
        wb.save(output)
    else:
        wb.save(str(output))


def export_to_csv(schedule: Schedule, output: Union[str, Path, io.StringIO]) -> None:
    """Export assignments to CSV."""
    df = schedule.to_dataframe()
    if isinstance(output, io.StringIO):
        df.to_csv(output, index=False)
    else:
        df.to_csv(str(output), index=False)
