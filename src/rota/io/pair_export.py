"""
Excel/CSV Export for Pair-Based Schedules
==========================================
Exports PairSchedule to Excel with all 6 legacy sheets and CSV with pair format.
"""
import io
from pathlib import Path
from typing import Any, Dict, List, Optional, Union

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from rota.models.person import Person
from rota.solver.edo import EDOPlan
from rota.solver.pairs import PairSchedule
from rota.solver.staffing import JOURS
from rota.solver.stats import calculate_person_stats
from rota.solver.validation import FairnessMetrics, ValidationResult
from rota.solver.weekend import WeekendResult
from rota.utils.logging_setup import get_logger

# Import shared constants and utils
from rota.io.pair_export_utils import (
    CODE, COLORS, THIN, DOUBLE_BLACK, BLACK_THIN, BORDER_THIN,
    get_shift_color
)

logger = get_logger("rota.io.pair_export")


def export_pairs_to_csv(
    schedule: PairSchedule,
    output: Union[str, Path, io.StringIO],
    days: List[str] = JOURS,
) -> None:
    """
    Export schedule to CSV with pair format.
    
    Columns: Semaine, Jour, Poste, Pair, Pers_A, Pers_B
    """
    rows = []
    for a in schedule.assignments:
        rows.append({
            "Semaine": a.week,
            "Jour": a.day,
            "Poste": a.shift,
            "Pair": a.slot_idx + 1,
            "Pers_A": a.person_a or "",
            "Pers_B": a.person_b or "",
        })
    
    df = pd.DataFrame(rows)
    
    # Sort
    day_order = {d: i for i, d in enumerate(days)}
    df["_day_order"] = df["Jour"].map(day_order)
    df = df.sort_values(["Semaine", "_day_order", "Poste", "Pair"])
    df = df.drop("_day_order", axis=1)
    
    if isinstance(output, (str, Path)):
        df.to_csv(output, index=False)
    else:
        df.to_csv(output, index=False)
    
    logger.info(f"Exported {len(rows)} assignments to CSV")


def export_pairs_to_excel(
    schedule: PairSchedule,
    people: List[Person],
    edo_plan: EDOPlan,
    output: Union[str, Path, io.BytesIO],
    validation: Optional[ValidationResult] = None,
    fairness: Optional[FairnessMetrics] = None,
    config: Optional[Dict[str, Any]] = None,
    staffing: Optional[Dict] = None,  # For gaps calculation
    days: List[str] = JOURS,
    team_borders: bool = False,
) -> None:
    """
    Export schedule to Excel with all sheets.
    
    Sheets:
        1. Tableau de bord - KPIs and options
        2. Matrice - Person × day grid
        3. ParPoste_Statique - Pairs per shift per day
        4. Synthèse - Per-person totals
        5. Technique - Validation details
        6. Violations - Detailed violations (if any)
        7. Gaps - Unfilled slots for contractors (if any)
    """
    logger.info(f"Exporting {len(schedule.assignments)} assignments to Excel")
    
    weeks = schedule.weeks
    config = config or {}
    
    # Build matrix data
    name_to_person = {p.name: p for p in people}
    names = sorted([p.name for p in people])
    
    # Build who works when using PairSchedule method
    works_on = schedule.get_person_day_matrix(code_map=CODE)
    
    # Fill in OFF and EDO
    for name in names:
        for w in range(1, weeks + 1):
            has_edo = name in edo_plan.plan.get(w, set())
            edo_day = edo_plan.fixed.get(name, "")
            edo_assigned = False
            
            for d in days:
                key = (name, w, d)
                if key not in works_on:
                    # Not working - mark as OFF or EDO
                    if has_edo and not edo_assigned:
                        if (not edo_day) or (edo_day == d):
                            works_on[key] = "EDO"
                            edo_assigned = True
                        else:
                            works_on[key] = "OFF"
                    else:
                        works_on[key] = "OFF"
                elif has_edo and not edo_assigned:
                    # Working but still need to assign EDO
                    pass  # Will get OFF or EDO* in remaining days
    
    # Create workbook
    wb = Workbook()
    
    # ========== Sheet 1: Tableau de bord ==========
    ws_db = wb.active
    ws_db.title = "Tableau de bord"
    
    # KPIs
    kpis = [
        ("Effectif", schedule.people_count),
        ("Semaines", weeks),
        ("Tous les postes pourvus", "Oui" if (validation is None or validation.slots_vides == 0) else "Non"),
        ("Score", f"{schedule.score:.2f}" if schedule.score else "—"),
    ]
    if validation:
        kpis.extend([
            ("Violations Nuit→Travail", validation.nuit_suivie_travail),
            ("Transitions Soir→Jour", validation.soir_vers_jour),
            ("Écarts hebdo", validation.ecarts_hebdo_jours),
            ("Écarts horizon", validation.ecarts_horizon_personnes),
        ])
    
    ws_db.cell(row=1, column=1, value="Indicateur").font = Font(bold=True)
    ws_db.cell(row=1, column=2, value="Valeur").font = Font(bold=True)
    for i, (label, val) in enumerate(kpis, start=2):
        ws_db.cell(row=i, column=1, value=label)
        ws_db.cell(row=i, column=2, value=val)
    
    # Options
    if config:
        row_start = len(kpis) + 4
        ws_db.cell(row=row_start-1, column=1, value="Options").font = Font(bold=True)
        for i, (k, v) in enumerate(config.items(), start=row_start):
            ws_db.cell(row=i, column=1, value=k)
            ws_db.cell(row=i, column=2, value=str(v))
    
    for i in range(1, 5):
        ws_db.column_dimensions[get_column_letter(i)].width = 25
    
    # ========== Sheet 2: Matrice ==========
    ws_m = wb.create_sheet("Matrice")
    
    # Headers
    _write_week_headers(ws_m, weeks, days)
    _write_days_row(ws_m, weeks, days)
    ws_m.freeze_panes = "B3"
    
    # Person rows
    for r, name in enumerate(names, start=3):
        ws_m.cell(row=r, column=1, value=name)
        c = 2
        for w in range(1, weeks + 1):
            for d in days:
                val = works_on.get((name, w, d), "OFF")
                cell = ws_m.cell(row=r, column=c, value=val)
                if val in COLORS:
                    cell.fill = PatternFill(start_color=COLORS[val], end_color=COLORS[val], fill_type="solid")
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = BORDER_THIN
                c += 1
    
    _apply_week_separators(ws_m, weeks, days)
    _write_col_widths(ws_m, 1 + weeks * len(days))
    
    # ========== Sheet 3: ParPoste_Statique ==========
    ws_pp = wb.create_sheet("ParPoste_Statique")
    
    _write_week_headers(ws_pp, weeks, days)
    _write_days_row(ws_pp, weeks, days)
    ws_pp.freeze_panes = "B3"
    
    # Build pairs per shift per day
    pairs_by = {}  # {(week, day, shift): ["A / B", "C / D"]}
    for a in schedule.assignments:
        key = (a.week, a.day, a.shift)
        pairs_by.setdefault(key, [])
        if a.shift == "S":  # Solo shift (Soir)
            pairs_by[key].append(a.person_a or "")
        else:
            pairs_by[key].append(f"{a.person_a} / {a.person_b}")
    
    shift_rows = [("Jour", "D"), ("Soir", "S"), ("Nuit", "N")]  # No Admin
    for r_idx, (label, shift) in enumerate(shift_rows, start=3):
        ws_pp.cell(row=r_idx, column=1, value=label).font = Font(bold=True)
        c = 2
        for w in range(1, weeks + 1):
            for d in days:
                pairs_list = pairs_by.get((w, d, shift), [])
                val = "; ".join(p for p in pairs_list if p.strip())
                cell = ws_pp.cell(row=r_idx, column=c, value=val)
                cell.alignment = Alignment(horizontal="left", vertical="center")
                cell.border = BORDER_THIN
                c += 1
    
    _apply_week_separators(ws_pp, weeks, days)
    _write_col_widths(ws_pp, 1 + weeks * len(days), width=20)
    
    # ========== Sheet 4: Synthèse ==========
    ws_syn = wb.create_sheet("Synthèse")
    
    syn_headers = ["Nom", "Jours", "Soirs", "Nuits", "Admin", "Total_Jours", "Cible", "Écart", "Nb_EDO"]
    for j, h in enumerate(syn_headers, start=1):
        ws_syn.cell(row=1, column=j, value=h).font = Font(bold=True)
    
    # Use centralized stats calculation
    person_stats = calculate_person_stats(schedule, people, edo_plan)
    
    for i, ps in enumerate(sorted(person_stats, key=lambda x: x.name), start=2):
        ws_syn.cell(row=i, column=1, value=ps.name)
        ws_syn.cell(row=i, column=2, value=ps.jours)
        ws_syn.cell(row=i, column=3, value=ps.soirs)
        ws_syn.cell(row=i, column=4, value=ps.nuits)
        ws_syn.cell(row=i, column=5, value=ps.admin)
        ws_syn.cell(row=i, column=6, value=ps.total)
        ws_syn.cell(row=i, column=7, value=ps.target)
        ws_syn.cell(row=i, column=8, value=ps.delta)
        ws_syn.cell(row=i, column=9, value=ps.edo_weeks)
    
    for j in range(1, len(syn_headers) + 1):
        ws_syn.column_dimensions[get_column_letter(j)].width = 14
    ws_syn.freeze_panes = "A2"
    
    # ========== Sheet 5: Technique ==========
    ws_tech = wb.create_sheet("Technique")
    
    if validation:
        val_data = validation.as_dict()
        tech_headers = list(val_data.keys())
        for j, h in enumerate(tech_headers, start=1):
            ws_tech.cell(row=1, column=j, value=h).font = Font(bold=True)
        for j, h in enumerate(tech_headers, start=1):
            ws_tech.cell(row=2, column=j, value=val_data[h])
    
    if fairness:
        row = 5
        ws_tech.cell(row=row-1, column=1, value="Équité par cohorte").font = Font(bold=True)
        ws_tech.cell(row=row, column=1, value="Cohorte").font = Font(bold=True)
        ws_tech.cell(row=row, column=2, value="σ Nuits").font = Font(bold=True)
        ws_tech.cell(row=row, column=3, value="σ Soirs").font = Font(bold=True)
        for i, (cid, std_n) in enumerate(fairness.night_std_by_cohort.items(), start=row+1):
            std_e = fairness.eve_std_by_cohort.get(cid, 0.0)
            ws_tech.cell(row=i, column=1, value=cid)
            ws_tech.cell(row=i, column=2, value=f"{std_n:.2f}")
            ws_tech.cell(row=i, column=3, value=f"{std_e:.2f}")
    
    for j in range(1, 6):
        ws_tech.column_dimensions[get_column_letter(j)].width = 18
    
    # ========== Sheet 6: Violations ==========
    if validation and validation.violations:
        ws_viol = wb.create_sheet("Violations")
        viol_headers = ["Sévérité", "Type", "Semaine", "Jour", "Personne", "Message"]
        
        # Write header
        for j, h in enumerate(viol_headers, start=1):
            cell = ws_viol.cell(row=1, column=j, value=h)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="333333", end_color="333333", fill_type="solid")
            cell.font = Font(bold=True, color="FFFFFF")
        
        # Write violations
        for i, v in enumerate(validation.violations, start=2):
            ws_viol.cell(row=i, column=1, value=v.severity)
            ws_viol.cell(row=i, column=2, value=v.type)
            ws_viol.cell(row=i, column=3, value=v.week)
            ws_viol.cell(row=i, column=4, value=v.day)
            ws_viol.cell(row=i, column=5, value=v.person)
            ws_viol.cell(row=i, column=6, value=v.message)
            
            # Color by severity
            if v.severity == "critical":
                fill = PatternFill(start_color="FF4444", end_color="FF4444", fill_type="solid")
            elif v.severity == "warning":
                fill = PatternFill(start_color="FFCC00", end_color="FFCC00", fill_type="solid")
            else:
                fill = None
            
            if fill:
                for col in range(1, 7):
                    ws_viol.cell(row=i, column=col).fill = fill
        
        for j in range(1, 7):
            ws_viol.column_dimensions[get_column_letter(j)].width = 18
        ws_viol.freeze_panes = "A2"
    
    # ========== Sheet 7: Gaps (Unfilled Slots) ==========
    if staffing:
        gaps_data = []
        weeks = schedule.weeks
        
        for w in range(1, weeks + 1):
            for d in days:
                for s_code in ["D", "S", "N"]:
                    req_slots = staffing[w].slots[d].get(s_code, 0)
                    assigned = len([a for a in schedule.assignments if a.week == w and a.day == d and a.shift == s_code])
                    
                    if assigned < req_slots:
                        missing_slots = req_slots - assigned
                        people_needed = missing_slots * 2 if s_code in ["D", "N"] else missing_slots
                        shift_name = {"D": "Jour", "S": "Soir", "N": "Nuit"}.get(s_code, s_code)
                        
                        gaps_data.append({
                            "Semaine": w,
                            "Jour": d,
                            "Quart": shift_name,
                            "Créneaux manquants": missing_slots,
                            "Personnes à recruter": people_needed,
                        })
        
        if gaps_data:
            ws_gaps = wb.create_sheet("Gaps")
            headers = ["Semaine", "Jour", "Quart", "Créneaux manquants", "Personnes à recruter"]
            
            # Write header
            for j, h in enumerate(headers, start=1):
                cell = ws_gaps.cell(row=1, column=j, value=h)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="FF4444", end_color="FF4444", fill_type="solid")
                cell.font = Font(bold=True, color="FFFFFF")
            
            # Write data
            for i, row in enumerate(gaps_data, start=2):
                ws_gaps.cell(row=i, column=1, value=row["Semaine"])
                ws_gaps.cell(row=i, column=2, value=row["Jour"])
                ws_gaps.cell(row=i, column=3, value=row["Quart"])
                ws_gaps.cell(row=i, column=4, value=row["Créneaux manquants"])
                ws_gaps.cell(row=i, column=5, value=row["Personnes à recruter"])
            
            # Total row
            total_row = len(gaps_data) + 2
            ws_gaps.cell(row=total_row, column=1, value="TOTAL").font = Font(bold=True)
            ws_gaps.cell(row=total_row, column=5, value=sum(g["Personnes à recruter"] for g in gaps_data)).font = Font(bold=True)
            
            for j in range(1, 6):
                ws_gaps.column_dimensions[get_column_letter(j)].width = 20
            ws_gaps.freeze_panes = "A2"
    
    # ============================================================
    # --- REORDER SHEETS: Management tabs first ---
    # ============================================================
    sheet_names = [ws.title for ws in wb.worksheets]
    
    # Define priority order (management first)
    priority = ["Tableau de bord", "Matrice", "Synthèse"]
    technical = ["ParPoste_Statique", "Technique", "Gaps", "Violations"]
    
    # Get sheets in priority order
    ordered = []
    for name in priority:
        if name in sheet_names:
            ordered.append(name)
    
    # Add technical sheets at the end
    for name in technical:
        if name in sheet_names and name not in ordered:
            ordered.append(name)
    
    # Add any remaining sheets
    for name in sheet_names:
        if name not in ordered:
            ordered.append(name)
    
    # Reorder workbook sheets
    for i, name in enumerate(ordered):
        sheet = wb[name]
        wb.move_sheet(sheet, offset=i - wb.worksheets.index(sheet))
    
    # Remove default "Sheet" if exists
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]
    
    # Save
    if isinstance(output, (str, Path)):
        wb.save(output)
    else:
        wb.save(output)
    
    logger.info(f"Excel export complete: {len(schedule.assignments)} assignments")



# Helper functions
def _write_week_headers(ws, weeks: int, days: List[str], start_col: int = 2, row: int = 1):
    """Write week headers with merged cells."""
    c = start_col
    for w in range(1, weeks + 1):
        ws.merge_cells(start_row=row, start_column=c, end_row=row, end_column=c + len(days) - 1)
        ws.cell(row=row, column=c, value=f"SEMAINE {w}").font = Font(bold=True)
        ws.cell(row=row, column=c).alignment = Alignment(horizontal="center", vertical="center")
        c += len(days)


def _write_days_row(ws, weeks: int, days: List[str], start_col: int = 2, row: int = 2):
    """Write day headers for each week."""
    c = start_col
    for w in range(1, weeks + 1):
        for d in days:
            cell = ws.cell(row=row, column=c, value=d)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            c += 1


def _apply_week_separators(ws, weeks: int, days: List[str], start_col: int = 2):
    """Add double borders between weeks."""
    max_row = ws.max_row
    for w in range(1, weeks + 1):
        c_start = start_col + (w - 1) * len(days)
        c_end = c_start + len(days) - 1
        for r in range(1, max_row + 1):
            # Left border
            cell = ws.cell(row=r, column=c_start)
            cell.border = Border(
                left=DOUBLE_BLACK, top=cell.border.top, bottom=cell.border.bottom, right=cell.border.right
            )
            # Right border
            cell = ws.cell(row=r, column=c_end)
            cell.border = Border(
                right=DOUBLE_BLACK, top=cell.border.top, bottom=cell.border.bottom, left=cell.border.left
            )


def _write_col_widths(ws, total_cols: int, width: int = 14):
    """Set column widths."""
    for i in range(1, total_cols + 1):
        ws.column_dimensions[get_column_letter(i)].width = width


def export_weekend_to_excel(
    result: WeekendResult,
    people: List[Person],
    output: Union[str, Path, io.BytesIO],
    num_weeks: int
) -> None:
    """Export weekend schedule to Excel."""
    logger.info(f"Exporting weekend schedule to Excel for {num_weeks} weeks")
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Planning Week-end"
    
    # Header: Name, Total, 24h, S1_Sat, S1_Sun, S2_Sat, ...
    headers = ["Nom", "Total", "24h"]
    days = ["Sam", "Dim"]
    
    # Build columns based on weeks
    for w in range(1, num_weeks + 1):
        for d in days:
            headers.append(f"S{w}_{d}")
            
    # Write Headers
    for c, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="333333", end_color="333333", fill_type="solid")
        cell.font = Font(bold=True, color="FFFFFF")
        
    ws.freeze_panes = "B2"
    
    # Process data
    w_map = {}
    for a in result.assignments:
        # map "Sat"->"Sam", "Sun"->"Dim" for consistency if needed, but here we use English keys from solver
        # Solver uses "Sat", "Sun"
        key = (a.person.name, a.week, a.day) # day is Sat/Sun
        if key not in w_map:
            w_map[key] = []
        w_map[key].append(a.shift)

    eligible_people = sorted([p for p in people if p.available_weekends], key=lambda p: p.name)
    
    for r, p in enumerate(eligible_people, start=2):
        ws.cell(row=r, column=1, value=p.name).font = Font(bold=True)
        
        total_shifts = 0
        shifts_24h = 0
        
        col_idx = 4
        for w in range(1, num_weeks + 1):
            for d_solver in ["Sat", "Sun"]:
                shifts = w_map.get((p.name, w, d_solver), [])
                shifts.sort()
                val = "+".join(shifts)
                
                cell = ws.cell(row=r, column=col_idx, value=val)
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = BORDER_THIN
                
                # Colors
                if len(shifts) == 2: # D+N
                     cell.fill = PatternFill(start_color="FFCCAA", end_color="FFCCAA", fill_type="solid")
                     shifts_24h += 1
                elif "D" in shifts:
                     cell.fill = PatternFill(start_color=COLORS["J"], end_color=COLORS["J"], fill_type="solid")
                elif "N" in shifts:
                     cell.fill = PatternFill(start_color=COLORS["N"], end_color=COLORS["N"], fill_type="solid")
                
                total_shifts += len(shifts)
                col_idx += 1
                
        ws.cell(row=r, column=2, value=total_shifts)
        ws.cell(row=r, column=3, value=shifts_24h)

    # Auto width
    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 8
    ws.column_dimensions["C"].width = 8
    for i in range(4, col_idx):
        ws.column_dimensions[get_column_letter(i)].width = 12

    if isinstance(output, (str, Path)):
        wb.save(output)
    else:
        wb.save(output)
    
    logger.info("Weekend export complete.")


def export_merged_calendar(
    weekday_schedule: PairSchedule,
    weekend_result: WeekendResult,
    people: List[Person],
    edo_plan: EDOPlan,
    output: Union[str, Path, io.BytesIO],
    validation: Optional[ValidationResult] = None,
    fairness: Optional[FairnessMetrics] = None,
    staffing: Optional[Dict] = None,
    config: Optional[Dict[str, Any]] = None,
) -> None:
    """
    Export merged calendar with:
    1. 'Tableau de bord': KPIs and options
    2. 'Vue Manager': Global view (Wide format: Weeks as columns)
    3. 'Synthèse': Per-person totals with weekend
    4. 'Technique': Validation breakdown
    5. 'Perso_{Name}': Individual calendar tabs
    6. 'Week-end': Weekend assignments
    """
    logger.info("Exporting merged calendar (Enhanced)")
    
    weeks = weekday_schedule.weeks
    config = config or {}
    days_week = ["Lun", "Mar", "Mer", "Jeu", "Ven"]
    days_weekend = ["Sam", "Dim"]
    days_all = days_week + days_weekend
    
    wb = Workbook()
    # Remove default sheet
    default_ws = wb.active
    wb.remove(default_ws)
    
    # Styles
    fill_day = PatternFill(start_color="DDEEFF", end_color="DDEEFF", fill_type="solid")
    fill_night = PatternFill(start_color="E6CCFF", end_color="E6CCFF", fill_type="solid")
    fill_solo = PatternFill(start_color="FFDDAA", end_color="FFDDAA", fill_type="solid")
    fill_24h = PatternFill(start_color="FFCCAA", end_color="FFCCAA", fill_type="solid")
    fill_header = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    
    font_bold = Font(bold=True)
    font_white = Font(bold=True, color="FFFFFF")
    align_center = Alignment(horizontal="center", vertical="center")
    
    # Border styles
    grey_thin = Side(style='thin', color='CCCCCC')
    black_double = Side(style='double', color='000000')
    border_grey = Border(left=grey_thin, right=grey_thin, top=grey_thin, bottom=grey_thin)
    
    # ============================================================
    # --- 0. TABLEAU DE BORD (Dashboard) ---
    # ============================================================
    ws_tdb = wb.create_sheet("Tableau de bord")
    
    # Header row
    ws_tdb.append(["Indicateur", "Valeur"])
    for cell in ws_tdb[1]:
        cell.font = font_white
        cell.fill = fill_header
        cell.alignment = align_center
    
    # General KPIs
    ws_tdb.append(["Effectif", len(people)])
    ws_tdb.append(["Semaines", weeks])
    ws_tdb.append(["Score interne", f"{weekday_schedule.score:.2f}" if weekday_schedule.score else "N/A"])
    ws_tdb.append(["Temps résolution", f"{weekday_schedule.solve_time_seconds:.1f}s" if weekday_schedule.solve_time_seconds else "N/A"])
    
    # Validation section
    ws_tdb.append([])  # Empty row
    ws_tdb.append(["--- Validation ---", ""])
    ws_tdb[ws_tdb.max_row][0].font = font_bold
    
    if validation:
        ws_tdb.append(["Slots vides", validation.slots_vides])
        incomplete = sum(1 for v in validation.violations if v.type == "incomplete_pair")
        ws_tdb.append(["Paires incomplètes", incomplete])
        ws_tdb.append(["Violations Nuit→Travail", validation.nuit_suivie_travail])
        ws_tdb.append(["Transitions Soir→Jour", validation.soir_vers_jour])
        ws_tdb.append(["Violations 48h", validation.rolling_48h_violations])
    else:
        ws_tdb.append(["(Validation non disponible)", ""])
    
    # Capacity Analysis section
    if staffing and edo_plan:
        ws_tdb.append([])
        ws_tdb.append(["--- Analyse Capacité ---", ""])
        ws_tdb[ws_tdb.max_row][0].font = font_bold
        
        try:
            from rota.solver.capacity import calculate_capacity
            cap = calculate_capacity(weekday_schedule, people, staffing, edo_plan)
            
            ws_tdb.append(["Capacité équipe (jours)", cap.net_capacity])
            ws_tdb.append(["Besoins totaux (shifts)", cap.total_required_person_shifts])
            ws_tdb.append(["Affectés (shifts)", cap.total_assigned_person_shifts])
            ws_tdb.append(["Balance", cap.capacity_balance])
            ws_tdb.append(["Utilisation (%)", f"{cap.utilization_percent:.1f}%"])
            
            # Per-shift breakdown
            for shift, data in cap.by_shift.items():
                shift_name = {"D": "Jour", "S": "Soir", "N": "Nuit"}.get(shift, shift)
                gap = data["gap"]
                status = "OK" if gap <= 0 else f"Manque {gap}"
                ws_tdb.append([f"  {shift_name}: Requis/Affectés/Écart", f"{data['required']}/{data['assigned']}/{status}"])
            
            # Recommendation
            if cap.agents_needed > 0.5:
                ws_tdb.append(["⚠️ Agents supplémentaires requis", f"+{cap.agents_needed:.1f}"])
            elif cap.excess_agent_days > 10:
                ws_tdb.append(["✅ Marge disponible (jours-agent)", f"{cap.excess_agent_days:.0f}"])
        except Exception as e:
            ws_tdb.append(["(Erreur analyse capacité)", str(e)])
    
    # Weekend section (if enabled)
    if weekend_result and weekend_result.assignments:
        ws_tdb.append([])
        ws_tdb.append(["--- Week-end ---", ""])
        ws_tdb[ws_tdb.max_row][0].font = font_bold
        ws_tdb.append(["Assignations WE", len(weekend_result.assignments)])
        ws_tdb.append(["Statut WE", weekend_result.status])

    
    # Options section
    ws_tdb.append([])
    ws_tdb.append(["--- Options ---", ""])
    ws_tdb[ws_tdb.max_row][0].font = font_bold
    ws_tdb.append(["EDO activé", "Oui" if config.get("edo_enabled") else "Non"])
    ws_tdb.append(["Nuits max séquence", config.get("max_nights_sequence", "N/A")])
    ws_tdb.append(["Mode équité", config.get("fairness_mode", "N/A")])
    ws_tdb.append(["Seed", config.get("seed", "N/A")])
    
    # Column widths
    ws_tdb.column_dimensions["A"].width = 25
    ws_tdb.column_dimensions["B"].width = 20
    
    # ============================================================
    # --- 1. VUE MANAGER (Global) ---
    # ============================================================
    ws_mgr = wb.create_sheet("Vue Manager")
    
    # Build Map
    full_map = {}  # (person, week, day) -> code
    
    # Weekday
    for a in weekday_schedule.assignments:
        code = a.shift  # "D", "N", "S"
        if a.person_a:
            full_map[(a.person_a, a.week, a.day)] = code
        if a.person_b:
            full_map[(a.person_b, a.week, a.day)] = code
            
    # Weekend
    if weekend_result:
        for a in weekend_result.assignments:
            key = (a.person.name, a.week, a.day)
            existing = full_map.get(key, "")
            if existing:
                full_map[key] = f"{existing}+{a.shift}"
            else:
                full_map[key] = a.shift

    # Headers for Manager View
    ws_mgr.cell(row=1, column=1, value="Nom").font = font_bold
    ws_mgr.cell(row=2, column=1, value="Nom").font = font_bold
    
    col_idx = 2
    week_start_cols = []  # Track first column of each week for borders
    for w in range(1, weeks + 1):
        start_col = col_idx
        week_start_cols.append(start_col)
        end_col = col_idx + 6
        ws_mgr.merge_cells(start_row=1, start_column=start_col, end_row=1, end_column=end_col)
        cell = ws_mgr.cell(row=1, column=start_col, value=f"Semaine {w}")
        cell.alignment = align_center
        cell.font = font_bold
        
        for d in days_all:
            c = ws_mgr.cell(row=2, column=col_idx, value=d)
            c.alignment = align_center
            c.border = border_grey
            if d in days_weekend:
                 c.font = Font(bold=True, color="FF0000")
            col_idx += 1
            
    # Headers for Totals
    headers_totals = ["Total J", "Total N", "Total S", "Total WE"]
    for i, h in enumerate(headers_totals):
        c = ws_mgr.cell(row=2, column=col_idx + i, value=h)
        c.font = font_bold
        c.alignment = align_center
        c.border = border_grey
    # Merge row 1 for "Totaux"
    ws_mgr.merge_cells(start_row=1, start_column=col_idx, end_row=1, end_column=col_idx + 3)
    ws_mgr.cell(row=1, column=col_idx, value="Totaux Période").font = font_bold
    ws_mgr.cell(row=1, column=col_idx).alignment = align_center

    # Data Rows
    active_people = sorted([p.name for p in people if p.workdays_per_week > 0])
    row_idx = 3
    for name in active_people:
        ws_mgr.cell(row=row_idx, column=1, value=name).font = font_bold
        ws_mgr.cell(row=row_idx, column=1).border = border_grey
        
        current_col = 2
        count_j = 0
        count_n = 0
        count_s = 0
        count_we = 0  # weekends worked (at least one shift)
        
        for w in range(1, weeks + 1):
            worked_this_we = False
            for d_idx, d in enumerate(days_all):
                val = full_map.get((name, w, d), "")
                cell = ws_mgr.cell(row=row_idx, column=current_col, value=val)
                cell.alignment = align_center
                cell.border = border_grey
                
                # Week separator: black double border on left of first column of each week
                if current_col in week_start_cols:
                    cell.border = Border(
                        left=black_double,
                        right=grey_thin,
                        top=grey_thin,
                        bottom=grey_thin
                    )
                
                # Counters
                if "D" in val:
                    count_j += 1
                if "N" in val:
                    count_n += 1
                if "S" in val:
                    count_s += 1
                
                # Weekend check
                if d in days_weekend and val:
                    worked_this_we = True
                
                # Styling (fills)
                if val == "N":
                    cell.fill = fill_night
                elif val == "D":
                    if d in days_weekend:
                        cell.fill = fill_solo
                    else:
                        cell.fill = fill_day
                elif val == "S":
                    cell.fill = fill_solo
                elif "N" in val and "D" in val:
                    cell.fill = fill_24h
                elif "D+N" in val:
                    cell.fill = fill_24h
                
                current_col += 1
            
            if worked_this_we:
                count_we += 1
        
        # Write Totals with Formulas
        end_col_char = get_column_letter(current_col - 1)
        r = row_idx
        range_str = f"B{r}:{end_col_char}{r}"
        
        # Total J
        c = ws_mgr.cell(row=r, column=current_col, value=f'=COUNTIF({range_str}, "*D*")')
        c.alignment = align_center
        c.font = font_bold
        c.border = border_grey
        
        # Total N
        c = ws_mgr.cell(row=r, column=current_col + 1, value=f'=COUNTIF({range_str}, "*N*")')
        c.alignment = align_center
        c.font = font_bold
        c.border = border_grey
        
        # Total S
        c = ws_mgr.cell(row=r, column=current_col + 2, value=f'=COUNTIF({range_str}, "*S*")')
        c.alignment = align_center
        c.font = font_bold
        c.border = border_grey
        
        # Total WE (Shifts on Sat/Sun)
        # Usage of SUMPRODUCT to count shifts in columns corresponding to Sam/Dim (indices 5, 6 mod 7)
        # Checks for D, N, S in those columns
        we_formula = (
            f'=SUMPRODUCT((MOD(COLUMN({range_str})-2,7)>=5)*'
            f'(ISNUMBER(SEARCH("D",{range_str}))+ISNUMBER(SEARCH("N",{range_str}))+ISNUMBER(SEARCH("S",{range_str}))))'
        )
        c = ws_mgr.cell(row=r, column=current_col + 3, value=we_formula)
        c.alignment = align_center
        c.font = font_bold
        c.border = border_grey

        row_idx += 1

    ws_mgr.freeze_panes = "B3"
    
    # --- 2. Personal Tabs ---
    for name in active_people:
        # Excel sheet name limit 31 chars
        safe_name = name.replace("/", "-").replace("\\", "-")[:20] 
        ws_p = wb.create_sheet(f"Perso_{safe_name}")
        
        ws_p.cell(row=1, column=1, value=f"Planning: {name}").font = Font(size=14, bold=True)
        
        # Header
        ws_p.cell(row=3, column=1, value="Semaine").font = font_bold
        for i, d in enumerate(days_all):
            ws_p.cell(row=3, column=2+i, value=d).font = font_bold
            
        r_p = 4
        for w in range(1, weeks + 1):
            ws_p.cell(row=r_p, column=1, value=f"S{w}").font = font_bold
            for i, d in enumerate(days_all):
                val = full_map.get((name, w, d), "")
                c = ws_p.cell(row=r_p, column=2+i, value=val)
                c.alignment = align_center
                if val == "N": c.fill = fill_night
                elif val == "D": c.fill = fill_day
                elif val == "S": c.fill = fill_solo
                elif "N" in val: c.fill = fill_24h
            r_p += 1
            
        for col in range(1, 9):
            ws_p.column_dimensions[get_column_letter(col)].width = 12

    # ============================================================
    # --- 3. SYNTHÈSE (Per-Person Summary) ---
    # ============================================================
    ws_syn = wb.create_sheet("Synthèse")
    
    # Headers - include weekend if available
    has_weekend = weekend_result and weekend_result.assignments
    if has_weekend:
        # Match Web Dashboard structure:
        # Jours, Soirs, Nuits, Admin, Total (Sem), Cible, Écart, Nb_EDO, WE Shifts, Total+WE
        headers = ["Nom", "Jours", "Soirs", "Nuits", "Admin", "Total (Sem)", "Cible", "Écart", "Nb_EDO", "WE Shifts", "Total+WE"]
    else:
        headers = ["Nom", "Jours", "Soirs", "Nuits", "Admin", "Total", "Cible", "Écart", "Nb_EDO"]
    
    ws_syn.append(headers)
    for cell in ws_syn[1]:
        cell.font = font_white
        cell.fill = fill_header
        cell.alignment = align_center
    
    # Count shifts per person
    we_counts = {}  # person -> weekend shift count
    if has_weekend:
        for a in weekend_result.assignments:
            we_counts[a.person.name] = we_counts.get(a.person.name, 0) + 1
    
    for p in sorted(people, key=lambda x: x.name):
        if p.workdays_per_week == 0:
            continue
        
        # Count weekday shifts only (exclude weekend days)
        jours = sum(1 for k, v in full_map.items() if k[0] == p.name and "D" in v and k[2] not in days_weekend)
        soirs = sum(1 for k, v in full_map.items() if k[0] == p.name and "S" in v and k[2] not in days_weekend)
        nuits = sum(1 for k, v in full_map.items() if k[0] == p.name and "N" in v and k[2] not in days_weekend)
        admin = 0  # Not currently tracked
        we_shifts = we_counts.get(p.name, 0)
        
        # Total (Semaine) - used for target comparison
        total_sem = jours + soirs + nuits + admin
        
        # Total including WE - for global view
        total_we = total_sem + we_shifts if has_weekend else total_sem

        edo_weeks = sum(1 for w in range(1, weeks + 1) if p.name in edo_plan.plan.get(w, set()))
        cible = p.workdays_per_week * weeks - edo_weeks
        
        # Écart matches Weekday Target compliance (as requested)
        ecart = total_sem - cible
        
        if has_weekend:
            # Structure: Weekday details -> Weekday Stats -> Weekend details -> Combined Total
            row = [p.name, jours, soirs, nuits, admin, total_sem, cible, ecart, edo_weeks, we_shifts, total_we]
        else:
            row = [p.name, jours, soirs, nuits, admin, total_sem, cible, ecart, edo_weeks]
        
        ws_syn.append(row)
        
        # Color-code écart: red for negative, green for positive
        # Color-code écart: red for negative, green for positive
        last_row = ws_syn.max_row
        ecart_col = 8 if has_weekend else 8 # Column H is index 8 in both cases now? 
        # Wait: 
        # Weekday: Nom(1), J(2), S(3), N(4), A(5), Tot(6), Cible(7), Ecart(8), EDO(9) -> Ecart is 8
        # Weekend: Nom(1), J(2), S(3), N(4), A(5), TotSem(6), Cible(7), Ecart(8), EDO(9), WE(10), TotWE(11) -> Ecart is 8
        # So Ecart is always column 8 ("H")
        ecart_col = 8
        if ecart < 0:
            ws_syn.cell(row=last_row, column=ecart_col).font = Font(color="FF0000", bold=True)
        elif ecart > 0:
            ws_syn.cell(row=last_row, column=ecart_col).font = Font(color="00AA00", bold=True)
    
    # Add summary row at bottom
    summary_row = ws_syn.max_row + 1
    ws_syn.cell(row=summary_row, column=1, value="TOTAL").font = Font(bold=True)
    
    # Sum each numeric column
    # Sum each numeric column
    if has_weekend:
        for col_idx in range(2, 12):  # Columns B through K (11 cols)
            # Numeric columns: J(2), S(3), N(4), A(5), TotSem(6), Cible(7), Ecart(8), EDO(9), WE(10), TotWE(11)
            # All are numeric except names
            col_letter = get_column_letter(col_idx)
            ws_syn.cell(row=summary_row, column=col_idx, 
                       value=f"=SUM({col_letter}2:{col_letter}{summary_row-1})")
            ws_syn.cell(row=summary_row, column=col_idx).font = Font(bold=True)
    else:
        for col_idx in range(2, 10):  # Columns B through I
            if col_idx <= 8:  # All numeric from 2 to 9 (EDO)
                col_letter = get_column_letter(col_idx)
                ws_syn.cell(row=summary_row, column=col_idx, 
                           value=f"=SUM({col_letter}2:{col_letter}{summary_row-1})")
            ws_syn.cell(row=summary_row, column=col_idx).font = Font(bold=True)
    
    # Style summary row
    fill_summary = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
    # Style summary row
    fill_summary = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
    for col_idx in range(1, (12 if has_weekend else 10)):
        ws_syn.cell(row=summary_row, column=col_idx).fill = fill_summary
    
    # Column widths
    ws_syn.column_dimensions["A"].width = 20
    
    # ============================================================
    # --- 4. TECHNIQUE (Validation Breakdown) ---
    # ============================================================
    ws_tech = wb.create_sheet("Technique")
    
    # Section 1: Validation counts
    ws_tech.append(["Métrique", "Valeur"])
    for cell in ws_tech[1]:
        cell.font = font_white
        cell.fill = fill_header
    
    if validation:
        ws_tech.append(["Slots_vides", validation.slots_vides])
        ws_tech.append(["doublons_jour", validation.doublons_jour])
        ws_tech.append(["Nuit_suivie_travail", validation.nuit_suivie_travail])
        ws_tech.append(["Soir_vers_Jour", validation.soir_vers_jour])
        ws_tech.append(["Ecarts_hebdo_jours", validation.ecarts_hebdo_jours])
        ws_tech.append(["Ecarts_horizon_personnes", validation.ecarts_horizon_personnes])
        ws_tech.append(["Violations_48h", validation.rolling_48h_violations])
    
    # Section 2: Cohort equity
    ws_tech.append([])
    ws_tech.append(["Équité par cohorte", ""])
    ws_tech[ws_tech.max_row][0].font = font_bold
    
    ws_tech.append(["Cohorte", "σ Nuits", "σ Soirs"])
    for cell in ws_tech[ws_tech.max_row]:
        cell.font = font_bold
    
    if fairness:
        for cohort, night_std in fairness.night_std_by_cohort.items():
            eve_std = fairness.eve_std_by_cohort.get(cohort, 0)
            ws_tech.append([cohort, f"{night_std:.2f}", f"{eve_std:.2f}"])
    
    # Column widths
    ws_tech.column_dimensions["A"].width = 25
    ws_tech.column_dimensions["B"].width = 15
    ws_tech.column_dimensions["C"].width = 15

    # --- 4. Gaps (Unfilled Shifts) ---
    # Include both unfilled_slot and incomplete_pair violations
    gap_types = {"unfilled_slot", "incomplete_pair"}
    gaps = [v for v in (validation.violations if validation else []) if v.type in gap_types]
    if gaps:
        ws_gaps = wb.create_sheet("Manques (Gaps)")
        ws_gaps.append(["Semaine", "Jour", "Poste", "Type", "Message"])
        for cell in ws_gaps[1]:
            cell.font = font_bold
        
        type_labels = {
            "unfilled_slot": "Slot vide",
            "incomplete_pair": "Paire incomplète"
        }
        for v in gaps:
            ws_gaps.append([v.week, v.day, v.shift, type_labels.get(v.type, v.type), v.message])
            
        ws_gaps.column_dimensions["E"].width = 50
    
    # --- 5. Weekend Sheet ---
    if weekend_result and weekend_result.assignments:
        ws_we = wb.create_sheet("Week-end")
        
        # Header
        ws_we.append(["Semaine", "Sam J", "Sam N", "Dim J", "Dim N"])
        for cell in ws_we[1]:
            cell.font = font_bold
            cell.alignment = align_center
        
        # Build weekend matrix by week
        we_data = {}  # {week: {day_shift: [names]}}
        for a in weekend_result.assignments:
            key = (a.week, a.day, a.shift)
            we_data.setdefault(key, []).append(a.person.name)
        
        for w in range(1, weeks + 1):
            sam_d = ", ".join(we_data.get((w, "Sam", "D"), ["-"]))
            sam_n = ", ".join(we_data.get((w, "Sam", "N"), ["-"]))
            dim_d = ", ".join(we_data.get((w, "Dim", "D"), ["-"]))
            dim_n = ", ".join(we_data.get((w, "Dim", "N"), ["-"]))
            ws_we.append([f"S{w}", sam_d, sam_n, dim_d, dim_n])
        
        # Column widths
        for col in ["B", "C", "D", "E"]:
            ws_we.column_dimensions[col].width = 30
    
    # --- 6. Violations Breakdown ---
    if validation and validation.violations:
        ws_viol = wb.create_sheet("Violations")
        ws_viol.append(["Type", "Sévérité", "Semaine", "Jour", "Poste", "Personne", "Message"])
        for cell in ws_viol[1]:
            cell.font = font_bold
        
        for v in validation.violations:
            ws_viol.append([v.type, v.severity, v.week, v.day, v.shift, v.person, v.message])
        
        ws_viol.column_dimensions["G"].width = 60
    
    # ============================================================
    # --- REORDER SHEETS: Management tabs first ---
    # ============================================================
    # Desired order: Tableau de bord, Vue Manager, Synthèse, [Perso tabs], Technique, Week-end, Manques, Violations
    sheet_names = [ws.title for ws in wb.worksheets]
    
    # Define priority order (management first)
    priority = ["Tableau de bord", "Vue Manager", "Synthèse"]
    technical = ["Technique", "Week-end", "Manques (Gaps)", "Violations"]
    
    # Get sheets in priority order
    ordered = []
    for name in priority:
        if name in sheet_names:
            ordered.append(name)
    
    # Add person sheets (Perso_*) in the middle
    for name in sheet_names:
        if name.startswith("Perso_") and name not in ordered:
            ordered.append(name)
    
    # Add technical sheets at the end
    for name in technical:
        if name in sheet_names and name not in ordered:
            ordered.append(name)
    
    # Add any remaining sheets
    for name in sheet_names:
        if name not in ordered:
            ordered.append(name)
    
    # Reorder workbook sheets
    for i, name in enumerate(ordered):
        sheet = wb[name]
        wb.move_sheet(sheet, offset=i - wb.worksheets.index(sheet))
    
    # Remove default "Sheet" if it exists and is empty
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]
    
    # Save
    if isinstance(output, (str, Path)):
        wb.save(output)
    else:
        wb.save(output)
    
    logger.info(f"Enhanced merged export complete: {len(active_people)} people, {weeks} weeks")
