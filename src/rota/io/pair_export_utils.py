"""
Shared Utilities for Pair Export
=================================
Constants, styles, and helper functions used by export modules.
"""
from openpyxl.styles import Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from rota.utils.logging_setup import get_logger

logger = get_logger("rota.io.pair_export_utils")

# Shift code map
CODE = {"D": "J", "E": "S", "N": "N", "S": "S"}  # E is legacy alias for S

# Colors (same as legacy)
COLORS = {
    "J": "DDEEFF",   # Day - light blue
    "S": "FFE4CC",   # Evening - light orange
    "N": "E6CCFF",   # Night - light purple
    "OFF": "EEEEEE", # Off - light grey
    "EDO": "D8D8D8", # EDO - darker grey
    "EDO*": "FFC7CE", # EDO conflict - light red
}

# Border styles
THIN = Side(border_style="thin", color="CCCCCC")
DOUBLE_BLACK = Side(border_style="double", color="000000")
BLACK_THIN = Side(border_style="thin", color="000000")
BORDER_THIN = Border(top=THIN, bottom=THIN, left=THIN, right=THIN)


def write_week_headers(ws, weeks: int, days: list, start_col: int = 2, row: int = 1):
    """Write week headers with merged cells."""
    for w in range(weeks):
        col_start = start_col + w * len(days)
        col_end = col_start + len(days) - 1
        
        ws.merge_cells(
            start_row=row, start_column=col_start,
            end_row=row, end_column=col_end
        )
        cell = ws.cell(row=row, column=col_start, value=f"Semaine {w + 1}")
        cell.font = Font(bold=True)
        cell.alignment = ws.cell(row=row, column=col_start).alignment


def write_days_row(ws, weeks: int, days: list, start_col: int = 2, row: int = 2):
    """Write day headers for each week."""
    for w in range(weeks):
        for i, d in enumerate(days):
            col = start_col + w * len(days) + i
            cell = ws.cell(row=row, column=col, value=d)
            cell.font = Font(bold=True)


def apply_week_separators(ws, weeks: int, days: list, start_col: int = 2):
    """Add double borders between weeks."""
    for w in range(1, weeks):
        col_idx = start_col + w * len(days)
        col_letter = get_column_letter(col_idx)
        for row in range(1, ws.max_row + 1):
            cell = ws[f"{col_letter}{row}"]
            existing = cell.border
            cell.border = Border(
                left=DOUBLE_BLACK,
                top=existing.top,
                bottom=existing.bottom,
                right=existing.right,
            )


def write_col_widths(ws, total_cols: int, width: int = 14):
    """Set column widths."""
    for c in range(1, total_cols + 1):
        ws.column_dimensions[get_column_letter(c)].width = width


def get_shift_color(shift: str) -> PatternFill:
    """Get fill color for a shift code."""
    color = COLORS.get(shift, COLORS.get("OFF", "FFFFFF"))
    return PatternFill(start_color=color, end_color=color, fill_type="solid")
