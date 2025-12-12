"""Tests for export functionality using mock data."""
import io
import pytest
from unittest.mock import MagicMock

from rota.models.person import Person
from rota.solver.pairs import PairAssignment, PairSchedule
from rota.solver.edo import EDOPlan
from rota.solver.validation import FairnessMetrics, ValidationResult
from rota.io.pair_export import export_pairs_to_csv, export_pairs_to_excel


@pytest.fixture
def mock_schedule():
    """Create a minimal mock schedule for testing exports."""
    assignments = [
        PairAssignment(week=1, day="Lun", shift="D", slot_idx=0, person_a="Alice", person_b="Bob"),
        PairAssignment(week=1, day="Lun", shift="N", slot_idx=0, person_a="Charlie", person_b="Diana"),
        PairAssignment(week=1, day="Mar", shift="D", slot_idx=0, person_a="Alice", person_b="Charlie"),
        PairAssignment(week=1, day="Mar", shift="S", slot_idx=0, person_a="Eve", person_b=""),
    ]
    return PairSchedule(
        assignments=assignments,
        weeks=1,
        people_count=5,
        status="optimal",
        score=0.0,
        solve_time_seconds=1.0,
    )


@pytest.fixture
def mock_people():
    """Create mock people list."""
    return [
        Person(name="Alice", workdays_per_week=4, edo_eligible=True),
        Person(name="Bob", workdays_per_week=4, edo_eligible=True),
        Person(name="Charlie", workdays_per_week=5, edo_eligible=False),
        Person(name="Diana", workdays_per_week=3, edo_eligible=True),
        Person(name="Eve", workdays_per_week=4, edo_eligible=True),
    ]


@pytest.fixture
def mock_edo_plan():
    """Create mock EDO plan."""
    return EDOPlan(plan={1: set()}, fixed={})


class TestCSVExport:
    """Tests for CSV export with mock data."""
    
    def test_export_to_csv_buffer(self, mock_schedule):
        """Test CSV export to StringIO buffer."""
        buffer = io.StringIO()
        export_pairs_to_csv(mock_schedule, buffer)
        
        content = buffer.getvalue()
        assert "Semaine" in content
        assert "Alice" in content
        assert "Bob" in content
        # Check CSV has proper structure
        lines = content.strip().split("\n")
        assert len(lines) >= 2  # Header + at least one row
        
    def test_export_csv_has_columns(self, mock_schedule):
        """Verify CSV has expected columns."""
        buffer = io.StringIO()
        export_pairs_to_csv(mock_schedule, buffer)
        
        content = buffer.getvalue()
        header = content.split("\n")[0]
        expected = ["Semaine", "Jour", "Poste"]
        for col in expected:
            assert col in header


class TestExcelExport:
    """Tests for Excel export with mock data."""
    
    def test_export_to_excel_buffer(self, mock_schedule, mock_people, mock_edo_plan):
        """Test Excel export to BytesIO buffer."""
        buffer = io.BytesIO()
        export_pairs_to_excel(
            schedule=mock_schedule,
            people=mock_people,
            edo_plan=mock_edo_plan,
            output=buffer,
        )
        
        # Check buffer has content (valid Excel file)
        assert len(buffer.getvalue()) > 0
        # Basic Excel file starts with PK signature
        buffer.seek(0)
        signature = buffer.read(2)
        assert signature == b"PK"
        
    def test_excel_has_sheets(self, mock_schedule, mock_people, mock_edo_plan, tmp_path):
        """Verify Excel has expected sheets."""
        from openpyxl import load_workbook
        
        path = tmp_path / "test_schedule.xlsx"
        export_pairs_to_excel(
            schedule=mock_schedule,
            people=mock_people,
            edo_plan=mock_edo_plan,
            output=path,
        )
        
        wb = load_workbook(path)
        sheet_names = wb.sheetnames
        
        # Check for key sheets
        assert "Tableau de bord" in sheet_names
        assert "Matrice" in sheet_names
        
    def test_excel_with_validation(self, mock_schedule, mock_people, mock_edo_plan):
        """Test Excel export with validation metrics."""
        buffer = io.BytesIO()
        
        validation = ValidationResult()
        fairness = FairnessMetrics(night_std=0.5, eve_std=0.3)
        
        export_pairs_to_excel(
            schedule=mock_schedule,
            people=mock_people,
            edo_plan=mock_edo_plan,
            output=buffer,
            validation=validation,
            fairness=fairness,
        )
        
        assert len(buffer.getvalue()) > 0
