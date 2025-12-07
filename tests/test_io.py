"""Tests for I/O functionality."""
import io
import tempfile
from pathlib import Path
import pytest
import pandas as pd

from rota.models.person import Person
from rota.models.schedule import Schedule, Assignment
from rota.models.shift import ShiftType, WEEKDAYS
from rota.io.csv_loader import load_team, save_team, team_to_dataframe
from rota.io.excel_export import export_to_excel, export_to_csv


class TestCSVLoader:
    """Tests for CSV loading functionality."""
    
    def test_load_team_from_file(self, team_dummy_path):
        """Test loading team from CSV file."""
        people = load_team(team_dummy_path)
        assert len(people) == 16
        assert all(isinstance(p, Person) for p in people)
    
    def test_load_team_from_dataframe(self):
        """Test loading team from DataFrame."""
        df = pd.DataFrame({
            "name": ["Alice", "Bob"],
            "workdays_per_week": [5, 4],
            "prefers_night": [0, 1],
        })
        people = load_team(df)
        assert len(people) == 2
        assert people[0].name == "Alice"
        assert people[0].workdays_per_week == 5
        assert people[1].prefers_night is True
    
    def test_load_team_missing_name_raises(self):
        """Test loading fails without name column."""
        df = pd.DataFrame({"workdays_per_week": [5, 4]})
        with pytest.raises(ValueError, match="name"):
            load_team(df)
    
    def test_load_team_handles_empty_values(self):
        """Test loader handles empty/missing values."""
        df = pd.DataFrame({
            "name": ["Test", ""],
            "workdays_per_week": [5, ""],
            "prefers_night": ["", "1"],
        })
        people = load_team(df)
        # Empty name row should be skipped
        assert len(people) == 1
        assert people[0].name == "Test"


class TestCSVSaver:
    """Tests for CSV saving functionality."""
    
    def test_save_team(self, sample_people, tmp_path):
        """Test saving team to CSV."""
        path = tmp_path / "team.csv"
        save_team(sample_people, path)
        
        assert path.exists()
        df = pd.read_csv(path)
        assert len(df) == len(sample_people)
        assert "name" in df.columns
    
    def test_save_empty_team(self, tmp_path):
        """Test saving empty team creates valid CSV."""
        path = tmp_path / "empty.csv"
        save_team([], path)
        
        assert path.exists()
        df = pd.read_csv(path)
        assert len(df) == 0
        assert "name" in df.columns
    
    def test_team_to_dataframe(self, sample_people):
        """Test converting team to DataFrame."""
        df = team_to_dataframe(sample_people)
        assert len(df) == len(sample_people)
        assert "name" in df.columns
        assert "workdays_per_week" in df.columns


class TestExcelExport:
    """Tests for Excel export functionality."""
    
    def test_export_to_excel_buffer(self, sample_people):
        """Test exporting to BytesIO buffer."""
        from rota.solver.engine import solve
        from rota.models.constraints import SolverConfig
        
        config = SolverConfig(weeks=2)
        schedule = solve(sample_people, config)
        
        buffer = io.BytesIO()
        export_to_excel(schedule, sample_people, buffer)
        
        assert len(buffer.getvalue()) > 0
    
    def test_export_to_excel_file(self, sample_people, tmp_path):
        """Test exporting to file."""
        from rota.solver.engine import solve
        from rota.models.constraints import SolverConfig
        
        config = SolverConfig(weeks=2)
        schedule = solve(sample_people, config)
        
        path = tmp_path / "schedule.xlsx"
        export_to_excel(schedule, sample_people, path)
        
        assert path.exists()
        assert path.stat().st_size > 0
    
    def test_excel_has_sheets(self, sample_people, tmp_path):
        """Test Excel file has expected sheets."""
        from rota.solver.engine import solve
        from rota.models.constraints import SolverConfig
        from openpyxl import load_workbook
        
        config = SolverConfig(weeks=2)
        schedule = solve(sample_people, config)
        
        path = tmp_path / "schedule.xlsx"
        export_to_excel(schedule, sample_people, path)
        
        wb = load_workbook(path)
        sheet_names = wb.sheetnames
        
        assert "Tableau de bord" in sheet_names
        assert "Matrice" in sheet_names
        assert "Synthèse" in sheet_names


class TestCSVExport:
    """Tests for CSV export functionality."""
    
    def test_export_to_csv_buffer(self, sample_people):
        """Test exporting to StringIO buffer."""
        from rota.solver.engine import solve
        from rota.models.constraints import SolverConfig
        
        config = SolverConfig(weeks=2)
        schedule = solve(sample_people, config)
        
        buffer = io.StringIO()
        export_to_csv(schedule, buffer)
        
        content = buffer.getvalue()
        assert "name" in content
        assert "week" in content
        assert "day" in content
        assert "shift" in content
    
    def test_export_to_csv_file(self, sample_people, tmp_path):
        """Test exporting to file."""
        from rota.solver.engine import solve
        from rota.models.constraints import SolverConfig
        
        config = SolverConfig(weeks=2)
        schedule = solve(sample_people, config)
        
        path = tmp_path / "schedule.csv"
        export_to_csv(schedule, path)
        
        assert path.exists()
        df = pd.read_csv(path)
        assert len(df) > 0
        assert set(df.columns) == {"name", "week", "day", "shift"}
